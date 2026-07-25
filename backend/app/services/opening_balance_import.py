"""Excel parser dùng chung cho import tồn đầu hàng loạt (kho NVL + kho thành phẩm) — chỉ đọc
và chuẩn hoá dữ liệu thô từng dòng theo mẫu 4 cột (Ngày nhập / Mã.../ Lô / Số lượng), KHÔNG tra
cứu vật tư/sản phẩm hay gọi receive()/build_units() — nghiệp vụ đó khác nhau giữa 2 nơi gọi
(services/warehouse.py và services/wms.py) nên để lại cho từng nơi tự xử lý."""

import io
from datetime import date, datetime

import openpyxl

from ..errors import DomainError


def _clean(v) -> str:
    return str(v).strip().upper() if v is not None else ""


def _parse_date_cell(val):
    """None/rỗng -> None (caller tự quyết định mặc định). Hỗ trợ ô Excel đã là ngày (openpyxl
    tự convert) hoặc chuỗi text YYYY-MM-DD / DD/MM/YYYY / DD-MM-YYYY."""
    if val in (None, ""):
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    raise ValueError(f"không đọc được ngày '{val}' (dùng YYYY-MM-DD hoặc DD/MM/YYYY)")


def parse_opening_balance_sheet(content: bytes, code_header: str, optional_headers: dict = None) -> list[dict]:
    """Đọc sheet đang active, tìm dòng tiêu đề (trong 10 dòng đầu) chứa đủ 4 cột bắt buộc:
    `code_header` (vd "MÃ VẬT TƯ" hoặc "MÃ SẢN PHẨM"), "NGÀY NHẬP", "LÔ", "SỐ LƯỢNG" (không
    phân biệt hoa thường/khoảng trắng thừa). `optional_headers` (vd {"vi_tri": "VỊ TRÍ"}) khai
    thêm các cột KHÔNG bắt buộc phải có trong sheet — nếu thiếu, giá trị tương ứng trả về None
    cho mọi dòng thay vì báo lỗi thiếu cột (file cũ theo đúng 4 cột gốc vẫn đọc được bình
    thường). Trả về list dict thô mỗi dòng dữ liệu theo đúng thứ tự trong Excel: {row, ma, lo,
    so_luong, ngay_nhap, error, **optional_headers}. `row` = số dòng thật trong sheet (để báo
    lỗi dễ đối chiếu lại file gốc). `error` (nếu có) mô tả lỗi đọc dữ liệu SỐ LƯỢNG/NGÀY NHẬP
    của riêng dòng đó — hàm này KHÔNG raise vì 1 dòng lỗi, để caller tự quyết định báo lỗi từng
    dòng mà vẫn tiếp tục xử lý các dòng còn lại."""
    optional_headers = optional_headers or {}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise DomainError(f"Không đọc được file Excel: {e}")
    ws = wb.active
    needed = [code_header, "NGÀY NHẬP", "LÔ", "SỐ LƯỢNG"]
    header_row, col_by_name = None, {}
    for r in range(1, min(ws.max_row, 10) + 1):
        found = {}
        for c in range(1, ws.max_column + 1):
            v = _clean(ws.cell(row=r, column=c).value)
            if v:
                found[v] = c
        if all(h in found for h in needed):
            header_row, col_by_name = r, found
            break
    if header_row is None:
        raise DomainError("Không tìm thấy dòng tiêu đề với đủ cột: " + ", ".join(needed) + ".")
    opt_cols = {key: col_by_name.get(hdr) for key, hdr in optional_headers.items()}

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        ma = ws.cell(row=r, column=col_by_name[code_header]).value
        lo = ws.cell(row=r, column=col_by_name["LÔ"]).value
        so_luong = ws.cell(row=r, column=col_by_name["SỐ LƯỢNG"]).value
        ngay = ws.cell(row=r, column=col_by_name["NGÀY NHẬP"]).value
        if ma in (None, "") and lo in (None, "") and so_luong in (None, "") and ngay in (None, ""):
            continue  # dòng trống — bỏ qua (không phải hết dữ liệu, có thể còn dòng sau)
        out = {"row": r, "ma": str(ma).strip() if ma is not None else "",
              "lo": str(lo).strip() if lo is not None else "", "so_luong": None,
              "ngay_nhap": None, "error": None}
        for key, col in opt_cols.items():
            val = ws.cell(row=r, column=col).value if col else None
            out[key] = str(val).strip() if val not in (None, "") else None
        if not out["ma"]:
            out["error"] = f"Thiếu {code_header.lower()}."
        try:
            out["so_luong"] = float(so_luong) if so_luong not in (None, "") else None
        except (TypeError, ValueError):
            out["so_luong"] = None
        if out["so_luong"] is None and not out["error"]:
            out["error"] = f"Số lượng không hợp lệ: '{so_luong}'."
        try:
            out["ngay_nhap"] = _parse_date_cell(ngay)
        except ValueError as e:
            if not out["error"]:
                out["error"] = f"Ngày nhập lỗi: {e}."
        rows.append(out)
    return rows
