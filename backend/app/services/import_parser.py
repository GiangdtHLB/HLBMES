"""Đọc file CSV/Excel → (columns, rows). Chỉ đọc, không import.

- CSV: stdlib csv, tự thử utf-8-sig (giữ tiếng Việt) rồi fallback.
- Excel: openpyxl, đọc sheet đầu, dòng đầu là header.
"""

import csv
import io
from typing import Tuple


def detect_source_type(filename: str) -> str:
    fn = (filename or "").lower()
    if fn.endswith(".xlsx") or fn.endswith(".xlsm"):
        return "excel"
    if fn.endswith(".csv") or fn.endswith(".txt"):
        return "csv"
    raise ValueError("Chỉ hỗ trợ .csv hoặc .xlsx")


def _norm(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return v


def parse_csv(data: bytes) -> Tuple[list, list]:
    # utf-8-sig để bỏ BOM + giữ tiếng Việt; fallback latin-1 nếu lỗi.
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("latin-1")
    reader = csv.reader(io.StringIO(text))
    rows_raw = [r for r in reader if any((c or "").strip() for c in r)]
    if not rows_raw:
        return [], []
    headers = [(_norm(h) or f"col{i}") for i, h in enumerate(rows_raw[0])]
    rows = []
    for r in rows_raw[1:]:
        row = {headers[i]: _norm(r[i]) if i < len(r) else "" for i in range(len(headers))}
        rows.append(row)
    return headers, rows


def parse_excel(data: bytes) -> Tuple[list, list]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header_row = next(it)
    except StopIteration:
        return [], []
    headers = [(_norm(h) or f"col{i}") for i, h in enumerate(header_row)]
    rows = []
    for r in it:
        if not any(c is not None and str(c).strip() for c in r):
            continue
        row = {headers[i]: _norm(r[i]) if i < len(r) else "" for i in range(len(headers))}
        rows.append(row)
    wb.close()
    return headers, rows


def parse(data: bytes, source_type: str) -> Tuple[list, list]:
    """Trả (columns, rows). rows là list[dict] theo header."""
    if source_type == "excel":
        return parse_excel(data)
    return parse_csv(data)
