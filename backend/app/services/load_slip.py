"""Nhập "Lệnh đóng hàng" (file Excel do bộ phận điều vận lập, 2 sheet HL/ĐM) và tách thành
các "Biên bản bàn giao hàng hóa" (LoadSlip) — mỗi xe (SỐ XE) trong sheet gộp thành 1 phiếu,
kèm danh mục hàng hóa (LoadSlipLine) lấy từ các cột SKU có số lượng > 0 trên các dòng của xe
đó. Đây là chứng từ giấy tờ nội bộ (Kho thành phẩm bàn giao cho xe/lái xe đi giao hàng),
KHÔNG trừ tồn kho WMS — khác với Shipment/Pallet vốn gắn với lô/pallet cụ thể.

Cột khuyến mại rời (LON/Lốc ... KM) đã được người lập lệnh tách riêng khỏi cột vỉ/thùng đủ
số ngay trong file gốc — ở đây chỉ cần giữ nguyên tách đó (is_promo=True, ĐVT Lon/Lốc), không
cần tính lại số lẻ khi 1 khuyến mại (VD 3 lon) không đủ đóng nguyên 1 vỉ 24 lon.

Ngoài ra, người lập lệnh đôi khi không tách cột riêng mà đánh dấu khuyến mại ngay trong 1 ô
của cột hàng chính (VD cột "Vỉ SP Sleek" bình thường không phải cột khuyến mại), theo 2 cách:
(a) gõ thẳng chữ "38 KM" vào ô — hiếm gặp; (b) CÁCH THẬT thường dùng: ô vẫn là SỐ bình thường
(38) nhưng định dạng số tùy chỉnh (Format Cells → Custom, VD `#,##0 "KM"`) để chỉ HIỂN THỊ
hậu tố "KM", không đổi kiểu dữ liệu — nhờ vậy các công thức tổng/SUM khác trong sheet không bị
hỏng. openpyxl đọc `.value` chỉ ra số thuần, phải đọc thêm `.number_format` mới bắt được cách
(b). Xem `_parse_cell_qty`. Cả 2 cách đều tách thành 1 dòng is_promo=True CÙNG tên sản phẩm
nhưng KHÔNG cộng gộp vào số lượng bán thường của cột đó.

File có thể khai báo thêm 1 dòng "Mã sản phẩm" ngay dưới dòng tiêu đề, gán mỗi cột hàng với
1 mã FinishedProduct.code (VD "CSPS330") — dùng để liên kết đúng SKU trong Danh mục Sản phẩm,
không chỉ dựa vào tên cột tự do. Cột khuyến mại rời "LON/Lốc ... KM" là hàng LON được phân rã
từ 1 cột vỉ/thùng nguyên (VD "Vỉ SP Sleek") — khai BẰNG ĐÚNG mã của cột gốc đó, không cần mã
riêng; is_promo=True + uom khác (Lon/Lốc thay vì Vỉ/Két) đã đủ phân biệt "cùng SKU, khác mức
đóng gói". File cũ không có dòng này thì product_code/finished_product_id đều để trống.

File cũng có thể khai báo thêm 1 dòng "Đơn vị tính" (khai TƯỜNG MINH đvt cho từng cột hàng,
VD "Vỉ", "Lon", "Két"...) — dòng này GHI ĐÈ đvt suy ra từ chữ đầu tên cột (_classify_column)
cho đúng cột nào có khai, cột nào không khai vẫn dùng đvt suy ra như cũ. Có thể xuất hiện cùng
lúc với dòng "Mã sản phẩm" nói trên, theo BẤT KỲ thứ tự nào (Mã sản phẩm trước rồi Đơn vị tính,
hoặc ngược lại) — cả 2 đều nằm liền nhau ngay dưới dòng tiêu đề, phân biệt bằng NỘI DUNG: dòng
mà đa số ô là nhãn đvt quen thuộc (Vỉ/Lon/Lốc/Chai/Két/Gòng/Lít/SL/Keg/Thùng) coi là dòng Đơn vị
tính; ngược lại (đa số là mã SKU dạng chữ+số) coi là dòng Mã sản phẩm. Xem _is_uom_label.

File có thể khai báo thêm 1 cột "MÃ NHÀ PHÂN PHỐI" (không bắt buộc, đặt cạnh cột "NPP VÀ NVBH")
— nếu có và khớp Supplier.code trong Danh mục Nhà cung cấp, TÊN nhà phân phối lấy từ danh mục
đó (Supplier.name) thay vì dùng chữ tự do ở cột "NPP VÀ NVBH"; mã không khớp hoặc cột không có
thì dùng lại đúng chữ tự do như cũ (không chặn import). Tương tự, SỐ XE luôn được tra trong Danh
mục xe (Vehicle.plate) — nếu khớp, TÊN LX lấy từ Vehicle.driver_name của danh mục thay vì chữ tự
do ở cột "TÊN LX"; xe chưa có trong danh mục thì vẫn dùng chữ tự do ở cột đó như cũ. Việc tra cứu
này cần truy vấn DB nên chỉ thực hiện trong import_casing_order (_parse_sheet thuần đọc file,
không đụng DB) — xem code_val/npp_codes bên dưới."""

import io
import re
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from ..common import new_id, utcnow
from ..errors import DomainError, NotFoundError
from ..models.master import FinishedProduct, UnitTypeCatalog
from ..models.materials import Supplier
from ..models.wms import LoadOrder, LoadSlip, LoadSlipLine, Shipment, Vehicle, WmsWarehouse

SHEET_TYPES = ["HL", "ĐM"]

# Các cột không phải hàng hóa (metadata điều vận/tài chính) — bỏ qua khi quét cột SKU.
_METADATA_HEADERS = {
    "pl", "cân tải trọng xe", "ghép tải ca 3", "ghép tải đm", "tổng trọng tải xe",
    "% trọng tải xe", "tiền hàng", "npp trả tm", "npp ck", "check hình ảnh",
    "số điện thoại lái xe", "thu vỏ keg", "thu vỏ chai", "thu pallet",
}
_STOP_MARKER = "tổng keg"
_CELL_KM_RE = re.compile(r"^\s*([\d]+(?:[.,]\d+)?)\s*km\.?\s*$", re.IGNORECASE)
# Nhãn đvt quen thuộc — dùng để phân biệt dòng "Đơn vị tính" với dòng "Mã sản phẩm" (mã SKU
# không bao giờ trùng các nhãn ngắn này).
_UOM_LABELS = {"vỉ", "vi", "lon", "lốc", "loc", "chai", "két", "ket", "gòng", "gông", "gong",
              "lít", "lit", "sl", "keg", "thùng", "thung"}


def _is_uom_label(txt: str) -> bool:
    return txt.strip().lower() in _UOM_LABELS


def _parse_cell_qty(cell) -> tuple[Optional[float], bool]:
    """Trả (qty, cell_is_promo). Nhận diện khuyến mại gõ thẳng vào 1 cột hàng chính (không có
    cột riêng) theo 2 cách, vì file thực tế dùng CẢ HAI tùy người lập:
    (a) ô CHỮ dạng "38 KM" — hiếm, ĐVT nằm ngay trong text.
    (b) ô SỐ bình thường (VD 38) nhưng định dạng số TÙY CHỈNH có hậu tố "KM" chỉ để HIỂN THỊ
        (Format Cells → Custom → #,##0 "KM") — CÁCH THẬT người lập lệnh hay dùng, vì giữ ô là
        số nên các công thức tổng/SUM khác trong sheet (Tổng keg, Tổng trọng tải xe...) không
        bị hỏng. openpyxl đọc .value là số thuần (38), không thấy chữ "KM" — phải kiểm tra
        thêm .number_format mới bắt được.
    Ô số bình thường không có gì đặc biệt -> (qty, False). Ô rỗng/không parse được -> (None, False)."""
    val = cell.value
    fmt_is_km = "km" in (cell.number_format or "").lower()
    if isinstance(val, str):
        m = _CELL_KM_RE.match(val)
        if m:
            return float(m.group(1).replace(",", ".")), True
        try:
            return float(val.replace(",", ".")), fmt_is_km
        except ValueError:
            return None, False
    try:
        return float(val), fmt_is_km
    except (TypeError, ValueError):
        return None, False


def _clean_header(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _classify_column(header: str) -> Optional[dict]:
    """Trả None nếu cột không phải hàng hóa; ngược lại {"uom", "is_promo"}."""
    h = _clean_header(header)
    if not h:
        return None
    hl = h.lower()
    if hl in _METADATA_HEADERS:
        return None
    is_promo = "km" in hl
    if hl.startswith("lon"):
        uom = "Lon"
    elif hl.startswith("lốc"):
        uom = "Lốc"
    elif hl.startswith("vỉ"):
        uom = "Vỉ"
    elif hl.startswith("gông"):
        uom = "Gông"
    elif hl.startswith("chai"):
        uom = "Chai" if hl.endswith("(chai)") else "Két"
    elif hl.startswith("bia hơi") or hl.startswith("bia tươi"):
        uom = "Lít"
    else:
        uom = "SL"
    return {"uom": uom, "is_promo": is_promo}


def _find_header_row(ws) -> int:
    for r in range(1, 16):
        for c in range(1, ws.max_column + 1):
            if _clean_header(ws.cell(row=r, column=c).value).upper() == "SỐ XE":
                return r
    raise DomainError("Không tìm thấy dòng tiêu đề (cột 'SỐ XE') trong sheet.")


def _parse_sheet_meta(ws) -> tuple[Optional[str], Optional[datetime]]:
    """Tìm 'Ca X' và 'Ngày D tháng M năm Y' trong vài dòng/ cột đầu sheet."""
    shift_label, order_date = None, None
    for r in range(1, 6):
        for c in range(1, 4):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            m_shift = re.search(r"Ca\s*\S+", v)
            if m_shift and not shift_label:
                shift_label = m_shift.group(0).strip()
            m_date = re.search(r"[Nn]gày\s*(\d+)\s*tháng\s*(\d+)\s*năm\s*(\d+)", v)
            if m_date and not order_date:
                d, mth, y = (int(x) for x in m_date.groups())
                try:
                    order_date = datetime(y, mth, d, tzinfo=timezone.utc)
                except ValueError:
                    pass
    return shift_label, order_date


def _parse_sheet(ws) -> list[dict]:
    header_row = _find_header_row(ws)
    col_by_name = {}
    for c in range(1, ws.max_column + 1):
        h = _clean_header(ws.cell(row=header_row, column=c).value).upper()
        if h and h not in col_by_name:
            col_by_name[h] = c
    required = ["SỐ XE", "TÊN LX", "NPP VÀ NVBH", "GHI CHÚ", "SỐ QĐ KM"]
    missing = [r for r in required if r not in col_by_name]
    if missing:
        raise DomainError(f"Sheet thiếu cột bắt buộc: {', '.join(missing)}.")
    c_plate, c_driver = col_by_name["SỐ XE"], col_by_name["TÊN LX"]
    c_npp, c_note, c_qdkm = col_by_name["NPP VÀ NVBH"], col_by_name["GHI CHÚ"], col_by_name["SỐ QĐ KM"]
    c_npp_code = col_by_name.get("MÃ NHÀ PHÂN PHỐI")  # cột tuỳ chọn — xem docstring đầu file

    product_cols = []  # [(col_idx, product_name, uom, is_promo)]
    for c in range(1, ws.max_column + 1):
        if c in (c_plate, c_driver, c_npp, c_note, c_qdkm, c_npp_code) or c in (
            col_by_name.get("TỔ LX"), col_by_name.get("TỔ NPP/NVBH")):
            continue
        header = _clean_header(ws.cell(row=header_row, column=c).value)
        cls = _classify_column(header)
        if cls:
            product_cols.append((c, header, cls["uom"], cls["is_promo"]))

    # Tối đa 2 dòng khai báo tuỳ chọn ngay dưới dòng tiêu đề — "Mã sản phẩm" (mỗi cột 1 mã
    # FinishedProduct.code, VD "CSPS330") và/hoặc "Đơn vị tính" (mỗi cột 1 đvt tường minh, VD
    # "Vỉ"), theo BẤT KỲ thứ tự nào. Nhận diện: cột SỐ XE trống + có ít nhất 1 ô hàng là chữ
    # không parse được thành số lượng; phân biệt loại dòng bằng nội dung — đa số ô khớp nhãn
    # đvt quen thuộc (_UOM_LABELS) => dòng Đơn vị tính, ngược lại => dòng Mã sản phẩm.
    code_by_col: dict[int, str] = {}
    uom_by_col: dict[int, str] = {}
    decl_rows: set[int] = set()
    r_scan = header_row + 1
    for _ in range(2):
        if r_scan > ws.max_row or _clean_header(ws.cell(row=r_scan, column=c_plate).value):
            break
        candidate: dict[int, str] = {}
        for c, name, uom, is_promo in product_cols:
            cell = ws.cell(row=r_scan, column=c)
            txt = _clean_header(cell.value)
            if not txt:
                continue
            qty, _ = _parse_cell_qty(cell)
            if qty is None:
                candidate[c] = txt
        if candidate:
            n_uom_like = sum(1 for t in candidate.values() if _is_uom_label(t))
            if n_uom_like * 2 >= len(candidate):
                uom_by_col = candidate
            else:
                code_by_col = candidate
            decl_rows.add(r_scan)
        r_scan += 1

    vehicles: list[dict] = []
    current = None
    for r in range(header_row + 1, ws.max_row + 1):
        if r in decl_rows:
            continue  # dòng khai Mã sản phẩm/Đơn vị tính — không phải dòng xe/hàng hóa
        a_val = _clean_header(ws.cell(row=r, column=c_plate).value)
        if a_val.lower() == _STOP_MARKER:
            break
        if a_val:
            if current is None or current["vehicle_plate"] != a_val:
                if current and current["lines"]:
                    vehicles.append(current)
                driver = _clean_header(ws.cell(row=r, column=c_driver).value) or None
                current = {"vehicle_plate": a_val, "driver_name": driver, "routes": [],
                          "npp_codes": [], "notes": [], "lines": {}}
        if current is None:
            continue  # dòng lẻ trước khi có xe đầu tiên — bỏ qua

        if c_npp_code:
            npp_code = _clean_header(ws.cell(row=r, column=c_npp_code).value)
            if npp_code and npp_code not in current["npp_codes"]:
                current["npp_codes"].append(npp_code)

        npp = _clean_header(ws.cell(row=r, column=c_npp).value)
        if npp and npp not in current["routes"]:
            current["routes"].append(npp)
        note = _clean_header(ws.cell(row=r, column=c_note).value)
        if note and note not in current["notes"] and note != "#N/A":
            current["notes"].append(note)
        qdkm = ws.cell(row=r, column=c_qdkm).value
        qdkm = str(qdkm).strip() if qdkm not in (None, "") else None

        for c, name, uom, col_is_promo in product_cols:
            cell = ws.cell(row=r, column=c)
            qty, cell_is_promo = _parse_cell_qty(cell)
            if not qty:
                continue
            is_promo = col_is_promo or cell_is_promo
            final_uom = uom_by_col.get(c, uom)
            key = (name, is_promo)
            line = current["lines"].setdefault(
                key, {"uom": final_uom, "is_promo": is_promo, "qty": 0.0, "qdkm": set(),
                      "product_name": name, "product_code": code_by_col.get(c)})
            line["qty"] += qty
            if is_promo and qdkm:
                line["qdkm"].add(qdkm)

    if current and current["lines"]:
        vehicles.append(current)
    return vehicles


def _next_slip_code(db: Session, year: int, used: dict) -> str:
    """Số thứ tự dựa trên SỐ LỚN NHẤT đã dùng cho năm đó (không phải COUNT) — vì xóa 1 phiếu
    ở giữa vẫn phải tiếp tục tăng từ số cao nhất, không được cấp lại trùng số đã xóa."""
    suffix = f"/{year}/BBBG-BHL"
    if year not in used:
        existing = db.execute(select(LoadSlip.slip_code)
                              .where(LoadSlip.slip_code.like(f"%{suffix}"))).scalars().all()
        max_seq = 0
        for code in existing:
            try:
                max_seq = max(max_seq, int(code.split("/")[0]))
            except (ValueError, IndexError):
                pass
        used[year] = max_seq
    used[year] += 1
    return f"{used[year]:03d}{suffix}"


def _next_order_code(db: Session, year: int, used: dict) -> str:
    """Số Lệnh đóng hàng — cùng quy tắc "số lớn nhất đã dùng" như _next_slip_code, nhưng
    đếm riêng theo hậu tố "/LDH-BHL" (Lệnh Đóng Hàng), không lẫn với số Biên bản bàn giao
    ("/BBBG-BHL")."""
    suffix = f"/{year}/LDH-BHL"
    if year not in used:
        existing = db.execute(select(LoadOrder.order_code)
                              .where(LoadOrder.order_code.like(f"%{suffix}"))).scalars().all()
        max_seq = 0
        for code in existing:
            try:
                max_seq = max(max_seq, int(code.split("/")[0]))
            except (ValueError, IndexError):
                pass
        used[year] = max_seq
    used[year] += 1
    return f"{used[year]:03d}{suffix}"


def _resolve_driver_name(vehicle_by_plate: dict, plate: str, sheet_driver: Optional[str]) -> Optional[str]:
    """SỐ XE khớp Danh mục xe (Vehicle.plate) -> ưu tiên Vehicle.driver_name; không khớp hoặc
    xe chưa có driver_name trong danh mục -> dùng lại chữ tự do ở cột TÊN LX như cũ."""
    vh = vehicle_by_plate.get(plate.strip().upper())
    if vh and vh.driver_name:
        return vh.driver_name
    return sheet_driver


def _resolve_routes(supplier_by_code: dict, npp_codes: list[str], sheet_routes: list[str]) -> list[str]:
    """Cột MÃ NHÀ PHÂN PHỐI (nếu có khai) là nguồn dữ liệu chính — mỗi mã khớp Supplier.code
    trả về đúng Supplier.name; mã không khớp trả lại chính mã đó (để lộ ra thay vì âm thầm bỏ
    qua, dễ phát hiện gõ sai). Không có cột này (hoặc không có mã nào trong xe) -> dùng lại
    danh sách tên NPP tự do ở cột NPP VÀ NVBH như cũ."""
    if not npp_codes:
        return sheet_routes
    resolved: list[str] = []
    for code in npp_codes:
        name = supplier_by_code.get(code.strip().upper()) or code
        if name not in resolved:
            resolved.append(name)
    return resolved


def import_casing_order(db: Session, filename: str, content: bytes, user,
                        sheet_type_overrides: dict[str, str] | None = None) -> dict:
    """sheet_type_overrides: {tên sheet thật trong file: "HL"|"ĐM"} — DỰ PHÒNG cho file có sheet
    KHÔNG đặt tên đúng "HL"/"ĐM" (quy ước chuẩn vẫn là đặt tên sheet đúng 2 mã này, tự nhận
    không cần khai gì thêm). Sheet nào không đúng tên chuẩn, không có trong overrides, NHƯNG
    thực sự có dữ liệu xe (parse được) sẽ KHÔNG import — trả về "needs_mapping" (tên các sheet
    đó) để người dùng chọn lại đúng sheet nào là HL/ĐM rồi gọi lại kèm overrides, thay vì đoán
    mò hoặc âm thầm bỏ qua dữ liệu thật."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise DomainError(f"Không đọc được file Excel: {e}")

    overrides = sheet_type_overrides or {}
    sheet_names_by_type: dict[str, list[str]] = {t: [] for t in SHEET_TYPES}
    needs_mapping: list[str] = []
    for name in wb.sheetnames:
        if name in SHEET_TYPES:
            sheet_names_by_type[name].append(name)
            continue
        mapped = overrides.get(name)
        if mapped in SHEET_TYPES:
            sheet_names_by_type[mapped].append(name)
            continue
        # Không đúng tên chuẩn, không có trong overrides — chỉ coi là "cần gán" nếu sheet này
        # THỰC SỰ có dữ liệu xe (parse được); sheet khác không liên quan (ghi chú/tham khảo...)
        # vẫn bỏ qua lặng lẽ như quy ước cũ.
        try:
            probe = _parse_sheet(wb[name])
        except DomainError:
            continue
        if probe:
            needs_mapping.append(name)
    if needs_mapping:
        return {"HL": [], "ĐM": [], "load_orders": {}, "needs_mapping": needs_mapping}

    fp_by_code = {fp.code.strip().upper(): fp.finished_product_id
                  for fp in db.execute(select(FinishedProduct)).scalars().all()}
    vehicle_by_plate = {v.plate.strip().upper(): v
                        for v in db.execute(select(Vehicle)).scalars().all() if v.plate}
    supplier_by_code = {s.code.strip().upper(): s.name
                        for s in db.execute(select(Supplier)).scalars().all() if s.code}

    seq_used: dict = {}
    order_seq_used: dict = {}
    created = {"HL": [], "ĐM": [], "load_orders": {}}
    for sheet_type, sheet_names in sheet_names_by_type.items():
        if not sheet_names:
            continue
        shift_label = order_date = None
        vehicles = []
        for name in sheet_names:
            ws = wb[name]
            sl, od = _parse_sheet_meta(ws)
            shift_label = shift_label or sl
            order_date = order_date or od
            vehicles.extend(_parse_sheet(ws))
        sheet_slips = []
        for v in vehicles:
            year = order_date.year if order_date else utcnow().year
            driver_name = _resolve_driver_name(vehicle_by_plate, v["vehicle_plate"], v["driver_name"])
            routes = _resolve_routes(supplier_by_code, v.get("npp_codes", []), v["routes"])
            slip = LoadSlip(
                load_slip_id=new_id(), slip_code=_next_slip_code(db, year, seq_used),
                sheet_type=sheet_type, shift_label=shift_label, order_date=order_date,
                vehicle_plate=v["vehicle_plate"], driver_name=driver_name,
                routes=", ".join(routes) or None, note=", ".join(v["notes"]) or None,
                source_file_name=filename, recipient_name=driver_name,
                recipient_unit=v["vehicle_plate"], created_by=user.username, created_at=utcnow(),
            )
            db.add(slip)
            db.flush()
            for i, line in enumerate(v["lines"].values()):
                note = f"Theo QĐ KM {', '.join(sorted(line['qdkm']))}" if line["qdkm"] else None
                code = line.get("product_code")
                db.add(LoadSlipLine(
                    line_id=new_id(), load_slip_id=slip.load_slip_id, seq=i,
                    product_name=line["product_name"], uom=line["uom"], quantity=round(line["qty"], 3),
                    is_promo=line["is_promo"], note=note, product_code=code,
                    finished_product_id=fp_by_code.get(code.strip().upper()) if code else None,
                ))
            sheet_slips.append(slip)
            created[sheet_type].append({"load_slip_id": slip.load_slip_id, "slip_code": slip.slip_code,
                                        "vehicle_plate": slip.vehicle_plate, "driver_name": slip.driver_name,
                                        "lines": len(v["lines"])})

        # 1 lần import = 1 "Lệnh đóng hàng" theo ngày cho MỖI sheet có xe — gộp tất cả xe của
        # sheet đó vào cùng 1 lệnh, dùng để in lại đúng layout bảng ngang của file Excel gốc.
        if sheet_slips:
            year = order_date.year if order_date else utcnow().year
            warehouse = _warehouse_for_sheet(db, sheet_type)
            order = LoadOrder(
                load_order_id=new_id(), order_code=_next_order_code(db, year, order_seq_used),
                sheet_type=sheet_type, shift_label=shift_label, order_date=order_date,
                source_file_name=filename, created_by=user.username, created_at=utcnow(),
                warehouse_id=warehouse.warehouse_id if warehouse else None,
            )
            db.add(order)
            db.flush()
            for slip in sheet_slips:
                slip.load_order_id = order.load_order_id
            created["load_orders"][sheet_type] = {"load_order_id": order.load_order_id,
                                                  "order_code": order.order_code}
    db.commit()
    return created


def list_load_slips(db: Session, sheet_type: Optional[str] = None,
                    limit: int = 1000, offset: int = 0) -> list[dict]:
    """Có phân trang (mặc định 1000, tối đa 5000) — số biên bản tích lũy tăng dần theo mỗi
    lần nhập lệnh đóng hàng. Đếm số dòng theo 1 truy vấn GROUP BY duy nhất thay vì 1 truy vấn
    COUNT riêng cho mỗi biên bản (N+1) như trước."""
    limit = max(1, min(limit or 1000, 5000))
    offset = max(0, offset or 0)
    q = select(LoadSlip).order_by(LoadSlip.created_at.desc()).limit(limit).offset(offset)
    if sheet_type:
        q = q.where(LoadSlip.sheet_type == sheet_type)
    slips = db.execute(q).scalars().all()
    slip_ids = [s.load_slip_id for s in slips]
    line_counts = dict(db.execute(
        select(LoadSlipLine.load_slip_id, func.count())
        .where(LoadSlipLine.load_slip_id.in_(slip_ids))
        .group_by(LoadSlipLine.load_slip_id)).all()) if slip_ids else {}
    shipment_ids = {s.shipment_id for s in slips if s.shipment_id}
    shipment_code_by_id = dict(db.execute(
        select(Shipment.shipment_id, Shipment.shipment_code)
        .where(Shipment.shipment_id.in_(shipment_ids))).all()) if shipment_ids else {}
    out = []
    for s in slips:
        n_lines = line_counts.get(s.load_slip_id, 0)
        out.append({
            "load_slip_id": s.load_slip_id, "slip_code": s.slip_code, "sheet_type": s.sheet_type,
            "shift_label": s.shift_label, "order_date": s.order_date, "vehicle_plate": s.vehicle_plate,
            "driver_name": s.driver_name, "routes": s.routes, "note": s.note,
            "recipient_name": s.recipient_name, "recipient_unit": s.recipient_unit,
            "created_at": s.created_at, "line_count": n_lines,
            "shipment_id": s.shipment_id, "shipment_code": shipment_code_by_id.get(s.shipment_id),
        })
    return out


def get_load_slip(db: Session, load_slip_id: str) -> dict:
    slip = db.get(LoadSlip, load_slip_id)
    if not slip:
        raise NotFoundError("Biên bản bàn giao hàng hóa không tồn tại.")
    lines = db.execute(select(LoadSlipLine).where(LoadSlipLine.load_slip_id == load_slip_id)
                       .order_by(LoadSlipLine.seq)).scalars().all()
    fp_ids = {l.finished_product_id for l in lines if l.finished_product_id}
    fp_by_id = {fp.finished_product_id: fp for fp in db.execute(
        select(FinishedProduct).where(FinishedProduct.finished_product_id.in_(fp_ids))
    ).scalars().all()} if fp_ids else {}
    shipment = db.get(Shipment, slip.shipment_id) if slip.shipment_id else None
    return {
        "load_slip_id": slip.load_slip_id, "slip_code": slip.slip_code, "sheet_type": slip.sheet_type,
        "shift_label": slip.shift_label, "order_date": slip.order_date, "vehicle_plate": slip.vehicle_plate,
        "driver_name": slip.driver_name, "routes": slip.routes, "note": slip.note,
        "source_file_name": slip.source_file_name,
        "issuer_name": slip.issuer_name, "issuer_title": slip.issuer_title, "issuer_dept": slip.issuer_dept,
        "recipient_name": slip.recipient_name, "recipient_title": slip.recipient_title,
        "recipient_unit": slip.recipient_unit, "created_by": slip.created_by, "created_at": slip.created_at,
        "shipment_id": slip.shipment_id, "shipment_code": shipment.shipment_code if shipment else None,
        "load_order_id": slip.load_order_id,
        "lines": [{"line_id": l.line_id, "seq": l.seq, "product_name": l.product_name, "uom": l.uom,
                   "quantity": l.quantity, "is_promo": l.is_promo, "note": l.note,
                   "product_code": l.product_code, "finished_product_id": l.finished_product_id,
                   "finished_product_name": fp_by_id[l.finished_product_id].name if l.finished_product_id in fp_by_id else None,
                   "finished_product_uom": fp_by_id[l.finished_product_id].uom if l.finished_product_id in fp_by_id else None,
                   } for l in lines],
    }


def update_load_slip_header(db: Session, load_slip_id: str, payload: dict) -> dict:
    slip = db.get(LoadSlip, load_slip_id)
    if not slip:
        raise NotFoundError("Biên bản bàn giao hàng hóa không tồn tại.")
    for key in ("issuer_name", "issuer_title", "issuer_dept",
                "recipient_name", "recipient_title", "recipient_unit"):
        if key in payload:
            setattr(slip, key, payload[key])
    db.commit()
    return get_load_slip(db, load_slip_id)


def update_load_slip_lines(db: Session, load_slip_id: str, lines: list[dict]) -> dict:
    """Sửa lại TOÀN BỘ dòng hàng (sản phẩm/SL/ĐVT/khuyến mại/ghi chú) của 1 xe — ghi đè hoàn
    toàn (xoá dòng cũ, tạo lại từ `lines`), cùng cách import_casing_order tạo dòng hàng loạt và
    delete_load_slip xoá hàng loạt, không có patch từng dòng riêng lẻ. CHỈ cho sửa khi xe CHƯA
    có phiếu xuất kho (shipment_id None) — xe đã xuất phải hoàn tác phiếu xuất đó trước (xem
    services/wms.py::undo_shipment, nơi mở lại shipment_id=None) mới sửa được tiếp."""
    slip = db.get(LoadSlip, load_slip_id)
    if not slip:
        raise NotFoundError("Biên bản bàn giao hàng hóa không tồn tại.")
    if slip.shipment_id:
        raise DomainError("Xe này đã có phiếu xuất kho — cần hoàn tác phiếu xuất đó trước khi sửa dòng hàng.")
    for l in db.execute(select(LoadSlipLine).where(LoadSlipLine.load_slip_id == load_slip_id)).scalars().all():
        db.delete(l)
    db.flush()
    for seq, l in enumerate(lines, start=1):
        product_name = (l.get("product_name") or "").strip()
        quantity = l.get("quantity") or 0
        if not product_name:
            raise DomainError(f"Dòng {seq}: thiếu tên sản phẩm.")
        if quantity <= 0:
            raise DomainError(f"Dòng {seq} ({product_name}): số lượng phải > 0.")
        db.add(LoadSlipLine(line_id=new_id(), load_slip_id=load_slip_id, seq=seq,
                            product_name=product_name, uom=(l.get("uom") or "").strip(),
                            quantity=quantity, is_promo=bool(l.get("is_promo")),
                            note=l.get("note") or None, product_code=l.get("product_code") or None,
                            finished_product_id=l.get("finished_product_id") or None))
    db.commit()
    return get_load_slip(db, load_slip_id)


def delete_load_slip(db: Session, load_slip_id: str) -> None:
    slip = db.get(LoadSlip, load_slip_id)
    if not slip:
        raise NotFoundError("Biên bản bàn giao hàng hóa không tồn tại.")
    for l in db.execute(select(LoadSlipLine).where(LoadSlipLine.load_slip_id == load_slip_id)).scalars().all():
        db.delete(l)
    db.delete(slip)
    db.commit()


# ---------------------------------------------------------------------------
# "Lệnh đóng hàng" theo ngày (LoadOrder) — gộp nhiều xe (LoadSlip) của cùng 1 lần import
# để in lại đúng layout bảng ngang của file Excel gốc. Xem docstring LoadOrder (models/wms.py).
# ---------------------------------------------------------------------------

def _order_warehouse(db: Session, order: LoadOrder) -> tuple[Optional[str], Optional[str]]:
    """Kho đóng hàng của 1 Lệnh đóng hàng — ưu tiên warehouse_id đã CHỐT lúc import (xem
    docstring cột LoadOrder.warehouse_id); nếu chưa có (import trước khi admin cấu hình
    WmsWarehouse.load_order_sheet_type khớp sheet_type), dò LIVE theo cấu hình hiện tại để
    admin cấu hình muộn vẫn dùng được ngay, không cần import lại."""
    if order.warehouse_id:
        wh = db.get(WmsWarehouse, order.warehouse_id)
        if wh:
            return wh.warehouse_id, wh.name
    wh = _warehouse_for_sheet(db, order.sheet_type)
    return (wh.warehouse_id, wh.name) if wh else (None, None)


def list_load_orders(db: Session, sheet_type: Optional[str] = None,
                     limit: int = 1000, offset: int = 0) -> list[dict]:
    limit = max(1, min(limit or 1000, 5000))
    offset = max(0, offset or 0)
    q = select(LoadOrder).order_by(LoadOrder.created_at.desc()).limit(limit).offset(offset)
    if sheet_type:
        q = q.where(LoadOrder.sheet_type == sheet_type)
    orders = db.execute(q).scalars().all()
    order_ids = [o.load_order_id for o in orders]
    vehicle_counts = dict(db.execute(
        select(LoadSlip.load_order_id, func.count())
        .where(LoadSlip.load_order_id.in_(order_ids))
        .group_by(LoadSlip.load_order_id)).all()) if order_ids else {}
    out = []
    for o in orders:
        warehouse_id, warehouse_name = _order_warehouse(db, o)
        out.append({
            "load_order_id": o.load_order_id, "order_code": o.order_code, "sheet_type": o.sheet_type,
            "shift_label": o.shift_label, "order_date": o.order_date, "source_file_name": o.source_file_name,
            "created_by": o.created_by, "created_at": o.created_at,
            "vehicle_count": vehicle_counts.get(o.load_order_id, 0),
            "warehouse_id": warehouse_id, "warehouse_name": warehouse_name,
        })
    return out


def get_load_order(db: Session, load_order_id: str) -> dict:
    order = db.get(LoadOrder, load_order_id)
    if not order:
        raise NotFoundError("Lệnh đóng hàng không tồn tại.")
    slips = db.execute(select(LoadSlip).where(LoadSlip.load_order_id == load_order_id)
                       .order_by(LoadSlip.vehicle_plate)).scalars().all()
    warehouse_id, warehouse_name = _order_warehouse(db, order)
    return {
        "load_order_id": order.load_order_id, "order_code": order.order_code, "sheet_type": order.sheet_type,
        "shift_label": order.shift_label, "order_date": order.order_date,
        "source_file_name": order.source_file_name, "created_by": order.created_by,
        "created_at": order.created_at,
        "warehouse_id": warehouse_id, "warehouse_name": warehouse_name,
        "vehicles": [get_load_slip(db, s.load_slip_id) for s in slips],
    }


def _norm_uom(txt) -> str:
    return re.sub(r"\s+", " ", str(txt or "")).strip().lower()


def _aggregate_export_lines(db: Session, vehicle_lines: list[tuple[dict, str]]) -> tuple[list[dict], list[dict]]:
    """Gộp danh sách dòng hàng (mỗi phần tử: (line dict kiểu get_load_slip()["lines"][i],
    vehicle_plate của xe chứa dòng đó)) thành gợi ý dòng Xuất kho — dùng chung cho
    export_lines_for_order (nhiều xe) và export_lines_for_slip (1 xe). CHỈ gộp các dòng khớp
    ĐÚNG mức đóng gói CHÍNH của SKU (VD "Vỉ Legend" khớp FinishedProduct.unit_type="vi" vì Danh
    mục Loại đơn vị tồn kho có code "vi" tên "Vỉ") — quantity ở mức này LUÔN là "số lượng theo
    unit_type" (xem services/wms.py::_pack_divisor), nên chuyển thẳng sang dòng Xuất kho không
    cần quy đổi gì. Dòng KHÔNG khớp (bán lẻ "Lon ... lẻ", "Chai ... lẻ", bia hơi tính theo Lít,
    hoặc chưa khớp được Mã sản phẩm) đưa vào "skipped" để thủ kho tự thêm tay qua picker có sẵn
    — không đoán/quy đổi sai unit_type."""
    fp_ids = {l["finished_product_id"] for l, _ in vehicle_lines if l["finished_product_id"]}
    fp_by_id = {fp.finished_product_id: fp for fp in db.execute(
        select(FinishedProduct).where(FinishedProduct.finished_product_id.in_(fp_ids))
    ).scalars().all()} if fp_ids else {}
    unit_type_name = {ut.code: ut.name for ut in db.execute(select(UnitTypeCatalog)).scalars().all()}

    agg: dict[tuple[str, bool], float] = {}
    skipped: list[dict] = []
    for l, vehicle_plate in vehicle_lines:
        fp = fp_by_id.get(l["finished_product_id"])
        if not fp or _norm_uom(l["uom"]) != _norm_uom(unit_type_name.get(fp.unit_type)):
            skipped.append({"product_name": l["product_name"], "uom": l["uom"],
                            "quantity": l["quantity"], "vehicle_plate": vehicle_plate})
            continue
        key = (fp.finished_product_id, l["is_promo"])
        agg[key] = agg.get(key, 0) + l["quantity"]

    lines = [{"product_name": fp_by_id[fpid].code, "unit_type": fp_by_id[fpid].unit_type,
             "quantity": round(qty, 3), "display_name": fp_by_id[fpid].name,
             "shipment_type": "promo" if is_promo else "normal"}
             for (fpid, is_promo), qty in agg.items()]
    return lines, skipped


def _warehouse_for_sheet(db: Session, sheet_type: str) -> Optional[WmsWarehouse]:
    return db.execute(select(WmsWarehouse)
                      .where(WmsWarehouse.load_order_sheet_type == sheet_type)).scalars().first()


def export_lines_for_order(db: Session, load_order_id: str) -> dict:
    """Gợi ý dòng Xuất kho từ 1 Lệnh đóng hàng, gộp TẤT CẢ xe trong đơn — xem
    _aggregate_export_lines cho logic khớp/quy đổi. warehouse_id/warehouse_name lấy từ
    LoadOrder.warehouse_id đã chốt lúc import (xem _order_warehouse)."""
    order = get_load_order(db, load_order_id)
    vehicle_lines = [(l, v["vehicle_plate"]) for v in order["vehicles"] for l in v["lines"]]
    lines, skipped = _aggregate_export_lines(db, vehicle_lines)
    return {"load_order_id": order["load_order_id"], "order_code": order["order_code"],
            "sheet_type": order["sheet_type"],
            "warehouse_id": order["warehouse_id"], "warehouse_name": order["warehouse_name"],
            "lines": lines, "skipped": skipped}


def export_lines_for_slip(db: Session, load_slip_id: str) -> dict:
    """Như export_lines_for_order nhưng CHỈ cho 1 xe cụ thể — dùng khi Xuất kho chọn đúng 1 xe
    trong Lệnh đóng hàng thay vì gộp cả đơn (xem views_ext.js, bảng chọn xe theo Sheet HL/ĐM).
    Kho lấy theo ĐÚNG Lệnh đóng hàng chứa xe này (LoadSlip.load_order_id) — không tự tra riêng
    theo sheet_type của xe, để nhất quán với export_lines_for_order (cùng 1 lệnh phải luôn ra
    cùng 1 kho dù xuất theo cả đơn hay theo từng xe). Xe không thuộc lệnh nào (hiếm — xem
    docstring LoadSlip.load_order_id) mới rơi về tra theo sheet_type riêng của xe đó."""
    slip = get_load_slip(db, load_slip_id)
    vehicle_lines = [(l, slip["vehicle_plate"]) for l in slip["lines"]]
    lines, skipped = _aggregate_export_lines(db, vehicle_lines)
    order_row = db.get(LoadOrder, slip["load_order_id"]) if slip["load_order_id"] else None
    if order_row:
        warehouse_id, warehouse_name = _order_warehouse(db, order_row)
    else:
        wh = _warehouse_for_sheet(db, slip["sheet_type"])
        warehouse_id, warehouse_name = (wh.warehouse_id, wh.name) if wh else (None, None)
    return {"load_slip_id": slip["load_slip_id"], "slip_code": slip["slip_code"],
            "vehicle_plate": slip["vehicle_plate"], "sheet_type": slip["sheet_type"],
            "warehouse_id": warehouse_id, "warehouse_name": warehouse_name,
            "lines": lines, "skipped": skipped}


def list_order_available_vehicles(db: Session, load_order_id: str) -> list[dict]:
    """Xe (LoadSlip) CHƯA thuộc lệnh đóng hàng nào, cùng sheet (HL/ĐM) với lệnh đang xem — dùng
    cho picker "+ Thêm xe". Xe đã ở lệnh này (hoặc lệnh khác) không hiện lại ở đây nữa."""
    order = db.get(LoadOrder, load_order_id)
    if not order:
        raise NotFoundError("Lệnh đóng hàng không tồn tại.")
    slips = db.execute(
        select(LoadSlip).where(LoadSlip.load_order_id.is_(None), LoadSlip.sheet_type == order.sheet_type)
        .order_by(LoadSlip.created_at.desc())
    ).scalars().all()
    return [{"load_slip_id": s.load_slip_id, "slip_code": s.slip_code, "vehicle_plate": s.vehicle_plate,
             "driver_name": s.driver_name} for s in slips]


def add_vehicle_to_order(db: Session, load_order_id: str, load_slip_id: str) -> dict:
    order = db.get(LoadOrder, load_order_id)
    if not order:
        raise NotFoundError("Lệnh đóng hàng không tồn tại.")
    slip = db.get(LoadSlip, load_slip_id)
    if not slip:
        raise NotFoundError("Biên bản bàn giao hàng hóa không tồn tại.")
    if slip.load_order_id:
        raise DomainError("Xe này đã thuộc 1 Lệnh đóng hàng khác — không thể thêm trùng.")
    if slip.sheet_type != order.sheet_type:
        raise DomainError(f"Xe thuộc sheet {slip.sheet_type}, không thể thêm vào lệnh sheet {order.sheet_type}.")
    slip.load_order_id = load_order_id
    db.commit()
    return get_load_order(db, load_order_id)


def remove_vehicle_from_order(db: Session, load_order_id: str, load_slip_id: str) -> dict:
    slip = db.get(LoadSlip, load_slip_id)
    if not slip or slip.load_order_id != load_order_id:
        raise NotFoundError("Xe này không thuộc Lệnh đóng hàng đang xem.")
    slip.load_order_id = None
    db.commit()
    return get_load_order(db, load_order_id)
