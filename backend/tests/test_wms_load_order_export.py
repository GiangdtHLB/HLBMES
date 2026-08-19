"""Test gợi ý dòng Xuất kho từ "Lệnh đóng hàng" (export_lines_for_order):
1) Dòng khớp ĐÚNG mức đóng gói chính của SKU (uom khớp tên UnitTypeCatalog của
   FinishedProduct.unit_type) -> gộp vào "lines", quantity giữ nguyên không quy đổi.
2) Dòng KHÔNG khớp (bán lẻ/chưa có mã) -> vào "skipped", không tự động thêm.
3) warehouse_id/warehouse_name suy đúng theo WmsWarehouse.load_order_sheet_type khớp
   order.sheet_type — None khi chưa gán kho nào.
4) PUT /wms/warehouses/{id} lưu đúng load_order_sheet_type."""

import io
import os
import tempfile

_TMP = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
os.environ["MES_DATABASE_URL"] = f"sqlite:///{_TMP.name}"
os.environ["MES_DEV_HEADER_AUTH"] = "0"
os.environ["MES_RL_ENABLED"] = "0"
os.environ["MES_ADMIN_PASSWORD"] = "AdminTest123"

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.main import app
from app import seed as seed_mod


@pytest.fixture(scope="module", autouse=True)
def _seeded():
    seed_mod.seed()
    yield


@pytest.fixture(scope="module")
def client():
    return TestClient(app)


def _login(client, u, p):
    r = client.post("/api/auth/login", json={"username": u, "password": p})
    assert r.status_code == 200, r.text
    return {"Authorization": "Bearer " + r.json()["token"]}


@pytest.fixture(scope="module")
def admin_h(client):
    return _login(client, "admin", "AdminTest123")


def _build_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HL"
    ws["A1"] = "LỆNH ĐÓNG HÀNG HL"
    ws["A2"] = "Ca 3  -  Ngày 18 tháng 8 năm 2026"
    headers = ["SỐ XE", "TÊN LX", "NPP VÀ NVBH", "GHI CHÚ", "SỐ QĐ KM", "Vỉ Legend", "Lon Legend lẻ"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=6, column=c, value=h)
    # Dòng khai "Mã sản phẩm" ngay dưới header — cả 2 cột cùng 1 SKU (loose lon là hàng phân rã
    # từ cùng SKU vỉ, xem docstring load_slip.py).
    ws.cell(row=7, column=6, value="CLGN330")
    ws.cell(row=7, column=7, value="CLGN330")
    # Xe duy nhất.
    ws.cell(row=8, column=1, value="11111")
    ws.cell(row=8, column=2, value="Tài xế A")
    ws.cell(row=8, column=5, value="NPP1")
    ws.cell(row=8, column=6, value=50)  # Vỉ Legend — khớp mức đóng gói chính (unit_type "vi")
    ws.cell(row=8, column=7, value=8)   # Lon Legend lẻ — KHÔNG khớp, phải vào skipped
    ws.cell(row=9, column=1, value="TỔNG KEG")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@pytest.fixture(scope="module")
def load_order_id(client, admin_h):
    fp = client.post("/api/finished-products", headers=admin_h, json={
        "code": "CLGN330", "name": "Bia lon Legend 330ml", "unit_type": "vi", "pack_size": 24,
    })
    assert fp.status_code == 201, fp.text
    content = _build_workbook()
    r = client.post("/api/wms/load-slips/import", headers=admin_h,
                    files={"file": ("lenh-dong-hang.xlsx", content,
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 201, r.text
    return r.json()["load_orders"]["HL"]["load_order_id"]


def test_case_level_line_aggregated_loose_line_skipped(client, admin_h, load_order_id):
    r = client.get(f"/api/wms/load-orders/{load_order_id}/export-suggestion", headers=admin_h)
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["sheet_type"] == "HL"
    assert len(data["lines"]) == 1
    line = data["lines"][0]
    assert line["product_name"] == "CLGN330"
    assert line["unit_type"] == "vi"
    assert line["quantity"] == 50
    assert line["shipment_type"] == "normal"
    assert len(data["skipped"]) == 1
    assert data["skipped"][0]["product_name"].strip() == "Lon Legend lẻ"
    assert data["skipped"][0]["quantity"] == 8


def test_warehouse_none_when_not_assigned(client, admin_h, load_order_id):
    r = client.get(f"/api/wms/load-orders/{load_order_id}/export-suggestion", headers=admin_h)
    assert r.status_code == 200, r.text
    assert r.json()["warehouse_id"] is None
    assert r.json()["warehouse_name"] is None


def test_warehouse_resolved_after_assignment_and_put_persists(client, admin_h, load_order_id):
    created = client.post("/api/wms/warehouses", headers=admin_h,
                          json={"code": "TEST-HL", "name": "Kho Test Hạ Long",
                                "load_order_sheet_type": "HL"})
    assert created.status_code == 201, created.text
    warehouse_id = created.json()["warehouse_id"]

    r = client.get(f"/api/wms/load-orders/{load_order_id}/export-suggestion", headers=admin_h)
    assert r.status_code == 200, r.text
    assert r.json()["warehouse_id"] == warehouse_id
    assert r.json()["warehouse_name"] == "Kho Test Hạ Long"

    updated = client.put(f"/api/wms/warehouses/{warehouse_id}", headers=admin_h,
                         json={"load_order_sheet_type": "ĐM"})
    assert updated.status_code == 200, updated.text
    listed = client.get("/api/wms/warehouses", headers=admin_h).json()
    row = next(w for w in listed if w["warehouse_id"] == warehouse_id)
    assert row["load_order_sheet_type"] == "ĐM"

    # Sau khi đổi sang "ĐM", lệnh sheet "HL" không còn khớp kho nào -> lại về None.
    r2 = client.get(f"/api/wms/load-orders/{load_order_id}/export-suggestion", headers=admin_h)
    assert r2.json()["warehouse_id"] is None
