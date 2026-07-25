"""Test import Excel hàng loạt tồn đầu (kho NVL công ty/phân xưởng + kho thành phẩm):
1) Admin-only ở cả 2 endpoint (thukho bị 403).
2) Dòng hợp lệ tạo đúng lô/đơn vị (kể cả để trống Lô -> tự sinh mã, ngày nhập xa quá 15 ngày
   vẫn được chấp nhận vì is_opening_balance).
3) Dòng lỗi (mã không tồn tại) không làm hỏng các dòng hợp lệ khác trong cùng file — trả về
   trong `failed`, không phải raise nguyên request.
4) Thiếu cột tiêu đề bắt buộc -> báo lỗi rõ ràng (409 DomainError), không phải lỗi khó hiểu."""

import io
import os
import tempfile

_TMP = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
os.environ["MES_DATABASE_URL"] = f"sqlite:///{_TMP.name}"
os.environ["MES_DEV_HEADER_AUTH"] = "0"
os.environ["MES_RL_ENABLED"] = "0"
os.environ["MES_ADMIN_PASSWORD"] = "AdminTest123"

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.main import app
from app import seed as seed_mod


@pytest.fixture(scope="module", autouse=True)
def _seeded():
    seed_mod.seed()
    yield


@pytest.fixture(scope="module")
def client():
    return TestClient(app)


def _login(client, u, p):
    r = client.post("/api/auth/login", json={"username": u, "password": p})
    assert r.status_code == 200, r.text
    return {"Authorization": "Bearer " + r.json()["token"]}


@pytest.fixture(scope="module")
def admin_h(client):
    return _login(client, "admin", "AdminTest123")


@pytest.fixture(scope="module")
def thukho_h(client):
    return _login(client, "thukho", "123456")


def _create_material(client, admin_h, code):
    r = client.post("/api/materials", headers=admin_h,
                    json={"code": code, "name": f"Vật tư {code}", "uom": "kg", "category": "other"})
    assert r.status_code == 201, r.text
    return r.json()["material_id"]


def _create_finished_product(client, admin_h, code):
    r = client.post("/api/finished-products", headers=admin_h,
                    json={"code": code, "name": f"SP {code}", "uom": "lon", "unit_type": "vi", "pack_size": 24})
    assert r.status_code == 201, r.text
    return r.json()


def _sheet(headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r_idx, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


_MAT_HEADERS = ["NGÀY NHẬP", "MÃ VẬT TƯ", "LÔ", "SỐ LƯỢNG"]
_FP_HEADERS = ["NGÀY NHẬP", "MÃ SẢN PHẨM", "LÔ", "SỐ LƯỢNG"]


def test_material_import_requires_admin(client, admin_h, thukho_h):
    mat_id = _create_material(client, admin_h, "OBX-NVL-A")
    content = _sheet(_MAT_HEADERS, [["2020-01-15", "OBX-NVL-A", "OBX-LOT-01", 100]])
    denied = client.post("/api/warehouse/opening-balance/import", headers=thukho_h,
                         data={"location": "Kho công ty"},
                         files={"file": ("ob.xlsx", content,
                                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert denied.status_code == 403, denied.text


def test_material_import_creates_lots_and_reports_row_errors(client, admin_h):
    mat_a = _create_material(client, admin_h, "OBX-NVL-B")
    content = _sheet(_MAT_HEADERS, [
        ["2020-01-15", "OBX-NVL-B", "OBX-LOT-02", 250],   # hợp lệ, ngày xa quá 15 ngày (được phép vì tồn đầu)
        ["2020-01-16", "OBX-NVL-B", "", 30],               # hợp lệ, lô để trống -> tự sinh mã
        ["2020-01-17", "OBX-NVL-KHONGTON", "OBX-LOT-03", 10],  # mã không tồn tại -> lỗi
        ["2020-01-18", "OBX-NVL-B", "OBX-LOT-04", "abc"],  # số lượng không hợp lệ -> lỗi
    ])
    r = client.post("/api/warehouse/opening-balance/import", headers=admin_h,
                    data={"location": "Kho phân xưởng"},
                    files={"file": ("ob.xlsx", content,
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    result = r.json()
    assert result["total"] == 4
    assert len(result["created"]) == 2
    assert len(result["failed"]) == 2
    reasons = " ".join(f["reason"] for f in result["failed"])
    assert "OBX-NVL-KHONGTON" in reasons
    assert "Số lượng" in reasons

    stock = client.get("/api/warehouse/stock", headers=admin_h, params={"location": "Kho phân xưởng"}).json()
    row = next(s for s in stock if s["material_id"] == mat_a)
    assert row["on_hand"] == 280  # 250 + 30


def test_units_import_requires_admin(client, admin_h, thukho_h):
    fp = _create_finished_product(client, admin_h, "OBX-FP-A")
    content = _sheet(_FP_HEADERS, [["2020-01-15", "OBX-FP-A", "OBX-FPLOT-01", 240]])
    denied = client.post("/api/wms/units/opening-balance/import", headers=thukho_h,
                         files={"file": ("ob.xlsx", content,
                                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert denied.status_code == 403, denied.text


def test_units_import_creates_units_and_reports_row_errors(client, admin_h):
    fp = _create_finished_product(client, admin_h, "OBX-FP-B")
    content = _sheet(_FP_HEADERS, [
        ["2020-01-15", "OBX-FP-B", "OBX-FPLOT-02", 240],
        ["2020-01-16", "OBX-FP-KHONGTON", "OBX-FPLOT-03", 24],
    ])
    r = client.post("/api/wms/units/opening-balance/import", headers=admin_h,
                    files={"file": ("ob.xlsx", content,
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 201, r.text
    result = r.json()
    assert result["total"] == 2
    assert len(result["created"]) == 1
    assert len(result["failed"]) == 1
    assert "OBX-FP-KHONGTON" in result["failed"][0]["reason"]

    units = client.get("/api/wms/units", headers=admin_h, params={"product": "OBX-FP-B"}).json()
    assert any(u.get("lot_code") == "OBX-FPLOT-02" for u in units)


def test_import_missing_header_returns_clear_error(client, admin_h):
    content = _sheet(["NGÀY NHẬP", "LÔ", "SỐ LƯỢNG"], [["2020-01-15", "OBX-LOT-05", 10]])
    r = client.post("/api/warehouse/opening-balance/import", headers=admin_h,
                    data={"location": "Kho công ty"},
                    files={"file": ("ob.xlsx", content,
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 409, r.text
    assert "MÃ VẬT TƯ" in r.json()["detail"]


def test_material_import_optional_kcs_lot_no_column(client, admin_h):
    """Cột SỐ LÔ KCS là tuỳ chọn — bỏ trống vẫn import bình thường, có giá trị thì lưu vào
    lô mới tạo (chỉ áp dụng khi TẠO lô mới, không đổi lô đã cộng dồn)."""
    mat_id = _create_material(client, admin_h, "OBX-NVL-KCS")
    content = _sheet(_MAT_HEADERS + ["SỐ LÔ KCS"], [
        ["2020-01-15", "OBX-NVL-KCS", "OBX-KCS-LOT-01", 100, "KCS-2020-001"],
        ["2020-01-16", "OBX-NVL-KCS", "OBX-KCS-LOT-02", 50, ""],
    ])
    r = client.post("/api/warehouse/opening-balance/import", headers=admin_h,
                    data={"location": "Kho công ty"},
                    files={"file": ("ob.xlsx", content,
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    result = r.json()
    assert len(result["created"]) == 2

    lots = client.get("/api/lots", headers=admin_h).json()
    lot1 = next(l for l in lots if l["lot_code"] == "OBX-KCS-LOT-01")
    lot2 = next(l for l in lots if l["lot_code"] == "OBX-KCS-LOT-02")
    assert lot1["kcs_lot_no"] == "KCS-2020-001"
    assert lot2["kcs_lot_no"] is None


def test_units_import_optional_location_column(client, admin_h):
    """Cột VỊ TRÍ là tuỳ chọn — mã hợp lệ thì gán vị trí ngay lúc tạo, mã không tồn tại báo lỗi
    dòng đó (không hỏng các dòng khác), để trống thì đơn vị ở trạng thái "chưa cất" như cũ."""
    fp = _create_finished_product(client, admin_h, "OBX-FP-LOC")
    loc = client.post("/api/wms/locations", headers=admin_h,
                      json={"code": "OBX-LOC-01", "name": "Kệ test vị trí", "capacity": 100})
    assert loc.status_code == 201, loc.text
    content = _sheet(_FP_HEADERS + ["VỊ TRÍ"], [
        ["2020-01-15", "OBX-FP-LOC", "OBX-FPLOC-LOT-01", 24, "OBX-LOC-01"],
        ["2020-01-16", "OBX-FP-LOC", "OBX-FPLOC-LOT-02", 24, "OBX-LOC-KHONGTON"],
        ["2020-01-17", "OBX-FP-LOC", "OBX-FPLOC-LOT-03", 24, ""],
    ])
    r = client.post("/api/wms/units/opening-balance/import", headers=admin_h,
                    files={"file": ("ob.xlsx", content,
                                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 201, r.text
    result = r.json()
    assert len(result["created"]) == 2
    assert len(result["failed"]) == 1
    assert "OBX-LOC-KHONGTON" in result["failed"][0]["reason"]

    units = client.get("/api/wms/units", headers=admin_h, params={"product": "OBX-FP-LOC"}).json()
    u1 = next(u for u in units if u["lot_code"] == "OBX-FPLOC-LOT-01")
    u3 = next(u for u in units if u["lot_code"] == "OBX-FPLOC-LOT-03")
    assert u1["location"] == "OBX-LOC-01"
    assert u3["location"] is None


def test_build_units_blocks_when_location_over_capacity(client, admin_h):
    fp = _create_finished_product(client, admin_h, "OBX-FP-CAP")
    loc = client.post("/api/wms/locations", headers=admin_h,
                      json={"code": "OBX-LOC-CAP", "name": "Kệ đầy", "capacity": 1})
    assert loc.status_code == 201, loc.text
    loc_id = loc.json()["loc_id"]
    ok = client.post("/api/wms/units", headers=admin_h,
                     json={"finished_product_id": fp["finished_product_id"], "product_name": fp["code"],
                           "lot_code": "OBX-CAP-LOT-01", "total": 24, "pack_size": 24, "unit_type": "vi",
                           "loc_id": loc_id})
    assert ok.status_code == 201, ok.text
    over = client.post("/api/wms/units", headers=admin_h,
                       json={"finished_product_id": fp["finished_product_id"], "product_name": fp["code"],
                             "lot_code": "OBX-CAP-LOT-02", "total": 24, "pack_size": 24, "unit_type": "vi",
                             "loc_id": loc_id})
    assert over.status_code == 409, over.text
    assert "đầy" in over.json()["detail"]
