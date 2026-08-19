"""Test 2 tính năng mới của import "Lệnh đóng hàng":
1) LoadOrder.warehouse_id được CHỐT ngay lúc import (theo WmsWarehouse.load_order_sheet_type
   khớp sheet_type tại thời điểm đó) — không tự đổi theo nếu sau đó admin cấu hình lại kho
   (khác hành vi cũ, dò LIVE mỗi lần đọc — vẫn giữ dò LIVE CHỈ khi warehouse_id còn None, xem
   test_wms_load_order_export.py::test_warehouse_resolved_after_assignment_and_put_persists).
2) Sheet không đặt tên đúng "HL"/"ĐM" mà có dữ liệu xe thật -> trả về "needs_mapping" thay vì
   âm thầm bỏ qua; import lại kèm sheet_type_overrides thì thành công."""

import io
import json
import os
import tempfile

_TMP = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
os.environ["MES_DATABASE_URL"] = f"sqlite:///{_TMP.name}"
os.environ["MES_DEV_HEADER_AUTH"] = "0"
os.environ["MES_RL_ENABLED"] = "0"
os.environ["MES_ADMIN_PASSWORD"] = "AdminTest123"

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.main import app
from app import seed as seed_mod


@pytest.fixture(scope="module", autouse=True)
def _seeded():
    seed_mod.seed()
    yield


@pytest.fixture(scope="module")
def client():
    return TestClient(app)


def _login(client, u, p):
    r = client.post("/api/auth/login", json={"username": u, "password": p})
    assert r.status_code == 200, r.text
    return {"Authorization": "Bearer " + r.json()["token"]}


@pytest.fixture(scope="module")
def admin_h(client):
    return _login(client, "admin", "AdminTest123")


def _build_workbook(sheet_name="HL"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws["A1"] = "LỆNH ĐÓNG HÀNG"
    ws["A2"] = "Ca 3  -  Ngày 19 tháng 8 năm 2026"
    headers = ["SỐ XE", "TÊN LX", "NPP VÀ NVBH", "GHI CHÚ", "SỐ QĐ KM", "Vỉ Legend"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=6, column=c, value=h)
    ws.cell(row=7, column=6, value="CLGN330IW")
    ws.cell(row=8, column=1, value="IW-01")
    ws.cell(row=8, column=2, value="Tài xế IW")
    ws.cell(row=8, column=5, value="NPP-IW")
    ws.cell(row=8, column=6, value=10)
    ws.cell(row=9, column=1, value="TỔNG KEG")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _import(client, admin_h, content, filename, overrides=None):
    data = {}
    if overrides is not None:
        data["sheet_type_overrides"] = json.dumps(overrides)
    return client.post("/api/wms/load-slips/import", headers=admin_h, data=data,
                       files={"file": (filename, content,
                                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})


@pytest.fixture(scope="module")
def finished_product(client, admin_h):
    fp = client.post("/api/finished-products", headers=admin_h, json={
        "code": "CLGN330IW", "name": "Bia lon Legend 330ml (import warehouse test)",
        "unit_type": "vi", "pack_size": 24,
    })
    assert fp.status_code == 201, fp.text
    return fp.json()


def test_warehouse_snapshotted_at_import_and_locked_after(client, admin_h, finished_product):
    wh = client.post("/api/wms/warehouses", headers=admin_h,
                     json={"code": "IW-HL", "name": "Kho Test IW Hạ Long", "load_order_sheet_type": "HL"})
    assert wh.status_code == 201, wh.text
    warehouse_id = wh.json()["warehouse_id"]

    content = _build_workbook("HL")
    r = _import(client, admin_h, content, "iw1.xlsx")
    assert r.status_code == 201, r.text
    load_order_id = r.json()["load_orders"]["HL"]["load_order_id"]

    order = client.get(f"/api/wms/load-orders/{load_order_id}", headers=admin_h).json()
    assert order["warehouse_id"] == warehouse_id
    assert order["warehouse_name"] == "Kho Test IW Hạ Long"

    listed = client.get("/api/wms/load-orders", headers=admin_h).json()
    row = next(o for o in listed if o["load_order_id"] == load_order_id)
    assert row["warehouse_id"] == warehouse_id

    exp = client.get(f"/api/wms/load-orders/{load_order_id}/export-suggestion", headers=admin_h).json()
    assert exp["warehouse_id"] == warehouse_id

    # Đổi cấu hình kho SAU khi đã import — đơn đã tạo phải GIỮ NGUYÊN kho đã chốt, không đổi
    # theo (khác hành vi "chưa từng chốt" — xem test_wms_load_order_export.py).
    updated = client.put(f"/api/wms/warehouses/{warehouse_id}", headers=admin_h,
                         json={"load_order_sheet_type": "ĐM"})
    assert updated.status_code == 200, updated.text

    order2 = client.get(f"/api/wms/load-orders/{load_order_id}", headers=admin_h).json()
    assert order2["warehouse_id"] == warehouse_id, "warehouse đã chốt lúc import không được đổi theo cấu hình mới"
    exp2 = client.get(f"/api/wms/load-orders/{load_order_id}/export-suggestion", headers=admin_h).json()
    assert exp2["warehouse_id"] == warehouse_id


def test_nonstandard_sheet_name_needs_mapping_then_succeeds_with_override(client, admin_h, finished_product):
    content = _build_workbook("Sheet1")
    r = _import(client, admin_h, content, "iw2.xlsx")
    assert r.status_code == 201, r.text
    body = r.json()
    assert body.get("needs_mapping") == ["Sheet1"]
    assert body["HL"] == [] and body["ĐM"] == []
    assert body["load_orders"] == {}

    r2 = _import(client, admin_h, content, "iw2.xlsx", overrides={"Sheet1": "HL"})
    assert r2.status_code == 201, r2.text
    body2 = r2.json()
    assert "needs_mapping" not in body2 or not body2["needs_mapping"]
    assert len(body2["HL"]) == 1
    load_order_id = body2["load_orders"]["HL"]["load_order_id"]
    order = client.get(f"/api/wms/load-orders/{load_order_id}", headers=admin_h).json()
    assert order["sheet_type"] == "HL"
    assert order["vehicles"][0]["vehicle_plate"] == "IW-01"


def test_irrelevant_sheet_without_vehicle_data_is_silently_ignored(client, admin_h, finished_product):
    # Sheet không đúng tên chuẩn NHƯNG cũng không parse được thành dữ liệu xe (VD sheet
    # ghi chú/tham khảo) -> vẫn bỏ qua lặng lẽ như quy ước cũ, không đòi hỏi gán.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GhiChu"
    ws["A1"] = "Chỉ là ghi chú nội bộ, không phải bảng lệnh đóng hàng."
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = _import(client, admin_h, buf.read(), "iw3.xlsx")
    assert r.status_code == 201, r.text
    body = r.json()
    assert not body.get("needs_mapping")
    assert body["HL"] == [] and body["ĐM"] == []
