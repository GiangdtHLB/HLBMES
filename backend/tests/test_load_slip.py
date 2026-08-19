"""Test nhập "Lệnh đóng hàng" (Excel, sheet HL/ĐM) → tách thành Biên bản bàn giao hàng hóa
theo xe (LoadSlip/LoadSlipLine):
1) Gộp đúng theo SỐ XE (kể cả dòng tiếp theo bỏ trống SỐ XE — kế thừa xe ngay trên).
2) Dòng "LON ... KM" (khuyến mại rời, chưa đủ 1 vỉ) tách thành dòng riêng is_promo=True,
   ĐVT "Lon" — không cộng gộp vào dòng "Vỉ" chính.
3) Dừng đúng tại mốc "TỔNG KEG" — không lấy nhầm dòng tổng/ghi chú/chữ ký ở cuối sheet.
4) CRUD qua API: import (multipart), list, get, update header (Bên giao/Bên nhận), delete."""

import io
import os
import tempfile

_TMP = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
os.environ["MES_DATABASE_URL"] = f"sqlite:///{_TMP.name}"
os.environ["MES_DEV_HEADER_AUTH"] = "0"
os.environ["MES_RL_ENABLED"] = "0"
os.environ["MES_ADMIN_PASSWORD"] = "AdminTest123"

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.main import app
from app import seed as seed_mod


@pytest.fixture(scope="module", autouse=True)
def _seeded():
    seed_mod.seed()
    yield


@pytest.fixture(scope="module")
def client():
    return TestClient(app)


def _login(client, u, p):
    r = client.post("/api/auth/login", json={"username": u, "password": p})
    assert r.status_code == 200, r.text
    return {"Authorization": "Bearer " + r.json()["token"]}


@pytest.fixture(scope="module")
def admin_h(client):
    return _login(client, "admin", "AdminTest123")


HEADERS = ["SỐ XE", "TÊN LX", "TỔ LX", "TỔ NPP/NVBH", "NPP VÀ NVBH", "GHI CHÚ", "SỐ QĐ KM",
          "Bia hơi 30L", "PL", "Vỉ Legend ", "LON Legend (Lon tết) KM"]


def _build_workbook(sheet_name="HL", shift="Ca 2", date_text="Ngày 5 tháng 3 năm 2026"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws["A1"] = f"LỆNH ĐÓNG HÀNG {sheet_name}"
    ws["A2"] = f"{shift}  -  {date_text}"
    for c, h in enumerate(HEADERS, start=1):
        ws.cell(row=6, column=c, value=h)
    # dòng rác ngay dưới header (giống file thật) — SỐ XE để trống, không được lấy nhầm.
    ws.cell(row=7, column=2, value=0)
    # Xe 11111: 1 dòng chính (Vỉ Legend=10) + 1 dòng tiếp (bỏ trống SỐ XE) có KM rời=3.
    ws.cell(row=8, column=1, value="11111")
    ws.cell(row=8, column=2, value="Tài xế A")
    ws.cell(row=8, column=5, value="NPP1")
    ws.cell(row=8, column=10, value=10)  # Vỉ Legend
    ws.cell(row=9, column=5, value="NPP1")
    ws.cell(row=9, column=7, value="100")  # SỐ QĐ KM
    ws.cell(row=9, column=11, value=3)  # LON Legend (Lon tết) KM
    # Xe 22222: 1 dòng, Bia hơi 30L=5.
    ws.cell(row=10, column=1, value="22222")
    ws.cell(row=10, column=2, value="Tài xế B")
    ws.cell(row=10, column=5, value="NPP2")
    ws.cell(row=10, column=8, value=5)  # Bia hơi 30L
    # Mốc dừng + rác phía sau — không được lấy vào danh sách xe.
    ws.cell(row=11, column=1, value="TỔNG KEG")
    ws.cell(row=11, column=8, value=15)
    ws.cell(row=12, column=1, value="99999")  # nếu parser không dừng đúng sẽ lẫn xe giả này
    ws.cell(row=12, column=2, value="Không nên xuất hiện")
    ws.cell(row=12, column=10, value=999)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _import(client, admin_h, content, filename="lenh-dong-hang.xlsx"):
    r = client.post("/api/wms/load-slips/import", headers=admin_h,
                    files={"file": (filename, content,
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 201, r.text
    return r.json()


def test_import_groups_by_vehicle_and_splits_promo_line(client, admin_h):
    content = _build_workbook()
    result = _import(client, admin_h, content)
    assert len(result["HL"]) == 2  # đúng 2 xe, không lẫn xe "99999" sau mốc TỔNG KEG
    assert result.get("ĐM", []) == []

    by_plate = {v["vehicle_plate"]: v for v in result["HL"]}
    assert by_plate["11111"]["lines"] == 2
    assert by_plate["22222"]["lines"] == 1

    slip = client.get(f"/api/wms/load-slips/{by_plate['11111']['load_slip_id']}", headers=admin_h).json()
    assert slip["driver_name"] == "Tài xế A"
    assert slip["routes"] == "NPP1"  # NPP lặp lại ở dòng tiếp theo không bị nhân đôi
    lines = {l["product_name"].strip(): l for l in slip["lines"]}
    assert lines["Vỉ Legend"]["quantity"] == 10
    assert lines["Vỉ Legend"]["uom"] == "Vỉ"
    assert lines["Vỉ Legend"]["is_promo"] is False
    promo = lines["LON Legend (Lon tết) KM"]
    assert promo["quantity"] == 3
    assert promo["uom"] == "Lon"
    assert promo["is_promo"] is True
    assert "100" in (promo["note"] or "")

    slip2 = client.get(f"/api/wms/load-slips/{by_plate['22222']['load_slip_id']}", headers=admin_h).json()
    assert slip2["lines"][0]["product_name"].strip() == "Bia hơi 30L"
    assert slip2["lines"][0]["uom"] == "Lít"


def test_cell_level_km_text_in_regular_column_splits_into_promo_line(client, admin_h):
    """Người lập lệnh đôi khi gõ thẳng "38 KM" vào ô của 1 cột hàng chính (không phải cột
    "...KM" riêng) — VD cột "Vỉ SP Sleek" của cùng 1 xe/NPP có nhiều dòng: dòng bán thường
    (208), rồi 2 dòng khuyến mại rời có số quyết định khác nhau (371, 366) và 1 dòng khuyến
    mại không có số quyết định. Phải tách đúng thành 2 dòng LoadSlipLine cho cùng sản phẩm:
    1 dòng thường (is_promo=False, chỉ có 208) và 1 dòng khuyến mại (is_promo=True, tổng
    38+8+6=52), ghi chú liệt kê đúng 2 số quyết định (không lẫn dòng không có số quyết định)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HL"
    ws["A1"] = "LỆNH ĐÓNG HÀNG HL"
    ws["A2"] = "Ca 3  -  Ngày 13 tháng 7 năm 2026"
    headers = ["SỐ XE", "TÊN LX", "NPP VÀ NVBH", "GHI CHÚ", "SỐ QĐ KM", "Vỉ SP Sleek"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=6, column=c, value=h)
    rows = [
        ("06631", "Tr Hà", "Nam Thu", None, 208),       # bán thường
        ("06631", None, "Nam Thu", None, "6 KM"),       # KM rời, không có số quyết định
        ("06631", None, "Nam Thu", "371", "38 KM"),     # KM rời, QĐ 371
        ("06631", None, "Nam Thu", "366", "8 KM"),      # KM rời, QĐ 366
    ]
    for i, (plate, driver, npp, qdkm, val) in enumerate(rows):
        r = 7 + i
        if plate:
            ws.cell(row=r, column=1, value=plate)
        if driver:
            ws.cell(row=r, column=2, value=driver)
        ws.cell(row=r, column=3, value=npp)
        if qdkm:
            ws.cell(row=r, column=5, value=qdkm)
        ws.cell(row=r, column=6, value=val)
    ws.cell(row=11, column=1, value="TỔNG KEG")
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    result = _import(client, admin_h, content)
    assert len(result["HL"]) == 1
    load_slip_id = result["HL"][0]["load_slip_id"]
    slip = client.get(f"/api/wms/load-slips/{load_slip_id}", headers=admin_h).json()
    lines = slip["lines"]
    assert len(lines) == 2, lines  # 1 dòng thường + 1 dòng khuyến mại gộp, không phải 4 dòng

    regular = next(l for l in lines if not l["is_promo"])
    promo = next(l for l in lines if l["is_promo"])
    assert regular["product_name"].strip() == "Vỉ SP Sleek"
    assert regular["quantity"] == 208
    assert regular["note"] is None

    assert promo["product_name"].strip() == "Vỉ SP Sleek"
    assert promo["quantity"] == 6 + 38 + 8
    assert "371" in promo["note"] and "366" in promo["note"]


def test_cell_level_km_custom_number_format_splits_into_promo_line(client, admin_h):
    """Cách THẬT người lập lệnh dùng trong file thực tế: ô vẫn là SỐ bình thường (VD 2), nhưng
    được định dạng số TÙY CHỈNH (Format Cells → Custom → #,##0 "KM") để chỉ HIỂN THỊ hậu tố
    "KM" — không đổi ô thành text — vì vậy các công thức SUM/tổng cộng khác trong sheet không
    bị hỏng. openpyxl đọc .value ra số thuần (2.0), KHÔNG thấy chữ "KM" trong value — phải đọc
    .number_format mới nhận diện được đây là dòng khuyến mại. Đây là biến thể QUAN TRỌNG NHẤT
    (khác với test gõ thẳng text "N KM" ở trên, vốn hiếm gặp trong thực tế)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HL"
    ws["A1"] = "LỆNH ĐÓNG HÀNG HL"
    ws["A2"] = "Ca 3  -  Ngày 13 tháng 7 năm 2026"
    headers = ["SỐ XE", "TÊN LX", "NPP VÀ NVBH", "GHI CHÚ", "SỐ QĐ KM", "Vỉ SP Sleek"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=6, column=c, value=h)
    # (plate, driver, npp, qdkm, value, is_km_format) — value luôn là SỐ, "KM" chỉ ở format.
    rows = [
        ("06631", "Tr Hà", "Nam Thu", None, 208, False),   # bán thường — format bình thường
        ("06631", None, "Nam Thu", None, 6, True),          # KM rời, không số quyết định
        ("06631", None, "Nam Thu", "371", 38, True),        # KM rời, QĐ 371
        ("06631", None, "Nam Thu", "366", 8, True),         # KM rời, QĐ 366
    ]
    for i, (plate, driver, npp, qdkm, val, is_km_fmt) in enumerate(rows):
        r = 7 + i
        if plate:
            ws.cell(row=r, column=1, value=plate)
        if driver:
            ws.cell(row=r, column=2, value=driver)
        ws.cell(row=r, column=3, value=npp)
        if qdkm:
            ws.cell(row=r, column=5, value=qdkm)
        cell = ws.cell(row=r, column=6, value=val)
        if is_km_fmt:
            cell.number_format = '#,##0 "KM"'
    ws.cell(row=11, column=1, value="TỔNG KEG")
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    result = _import(client, admin_h, content)
    assert len(result["HL"]) == 1
    load_slip_id = result["HL"][0]["load_slip_id"]
    slip = client.get(f"/api/wms/load-slips/{load_slip_id}", headers=admin_h).json()
    lines = slip["lines"]
    assert len(lines) == 2, lines

    regular = next(l for l in lines if not l["is_promo"])
    promo = next(l for l in lines if l["is_promo"])
    assert regular["quantity"] == 208
    assert regular["note"] is None

    assert promo["quantity"] == 6 + 38 + 8
    assert "371" in promo["note"] and "366" in promo["note"]


def test_product_code_row_links_finished_product_and_decomposed_lon_shares_code(client, admin_h):
    """File có thể khai thêm 1 dòng "Mã sản phẩm" ngay dưới header, gán mỗi cột hàng 1 mã SKU
    (FinishedProduct.code). Cột "LON ... KM" là hàng lon PHÂN RÃ từ cột "Vỉ SP Sleek" — khai
    CÙNG mã CSPS330 (không mã riêng) — is_promo + uom khác đã đủ phân biệt 2 dòng khác nhau
    của cùng 1 SKU. Mã không khớp FinishedProduct nào (VD "XX999") vẫn giữ product_code
    nhưng finished_product_id=None (không lỗi, chỉ không liên kết được)."""
    fp = client.post("/api/finished-products", headers=admin_h, json={
        "code": "CSPS330", "name": "Vỉ SP Sleek 330ml", "unit_type": "vi", "pack_size": 24,
    })
    assert fp.status_code == 201, fp.text

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HL"
    ws["A1"] = "LỆNH ĐÓNG HÀNG HL"
    ws["A2"] = "Ca 3  -  Ngày 13 tháng 7 năm 2026"
    headers = ["SỐ XE", "TÊN LX", "NPP VÀ NVBH", "GHI CHÚ", "SỐ QĐ KM", "Vỉ SP Sleek", "LON SP Sleek KM"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=6, column=c, value=h)
    # Dòng mã sản phẩm ngay dưới header — SỐ XE để trống. Mã "XX999" cố tình không khớp SKU
    # nào để kiểm tra vẫn import được, chỉ không liên kết finished_product_id.
    ws.cell(row=7, column=6, value="CSPS330")
    ws.cell(row=7, column=7, value="XX999")
    ws.cell(row=8, column=1, value="06631")
    ws.cell(row=8, column=2, value="Tr Hà")
    ws.cell(row=8, column=3, value="Nam Thu")
    ws.cell(row=8, column=6, value=208)
    ws.cell(row=9, column=1, value="06631")
    ws.cell(row=9, column=3, value="Nam Thu")
    ws.cell(row=9, column=7, value=16)
    ws.cell(row=10, column=1, value="TỔNG KEG")
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    result = _import(client, admin_h, content)
    slip = client.get(f"/api/wms/load-slips/{result['HL'][0]['load_slip_id']}", headers=admin_h).json()
    lines = {l["product_name"].strip(): l for l in slip["lines"]}

    regular = lines["Vỉ SP Sleek"]
    assert regular["product_code"] == "CSPS330"
    assert regular["finished_product_id"] == fp.json()["finished_product_id"]
    assert regular["finished_product_name"] == "Vỉ SP Sleek 330ml"
    assert regular["finished_product_uom"] == fp.json()["uom"]  # ĐVT chuẩn theo Danh mục Sản phẩm
    assert regular["is_promo"] is False

    promo = lines["LON SP Sleek KM"]
    assert promo["product_code"] == "XX999"
    assert promo["finished_product_id"] is None  # mã lạ, không khớp SKU nào
    assert promo["finished_product_uom"] is None  # không khớp -> không có ĐVT chuẩn
    assert promo["is_promo"] is True


def test_uom_row_overrides_header_inferred_uom_and_coexists_with_code_row(client, admin_h):
    """File có thể khai thêm 1 dòng "Đơn vị tính" (ngoài dòng "Mã sản phẩm" đã có) để khai
    THẲNG đvt cho từng cột, ghi đè đvt suy ra từ chữ đầu tên cột — VD cột "Vỉ SP Sleek" đáng lẽ
    tự suy ra "Vỉ", nhưng dòng Đơn vị tính khai "Thùng" thì phải lấy "Thùng". Cả 2 dòng khai báo
    (Mã sản phẩm ở dòng 7, Đơn vị tính ở dòng 8 — thứ tự này hay ngược lại đều phải nhận đúng)
    cùng tồn tại ngay dưới header, không được lẫn vào dữ liệu xe."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HL"
    ws["A1"] = "LỆNH ĐÓNG HÀNG HL"
    ws["A2"] = "Ca 3  -  Ngày 13 tháng 7 năm 2026"
    headers = ["SỐ XE", "TÊN LX", "NPP VÀ NVBH", "GHI CHÚ", "SỐ QĐ KM",
              "Vỉ SP Sleek", "LON SP Sleek KM", "Bia hơi 30L"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=6, column=c, value=h)
    # Dòng 7: Mã sản phẩm. Dòng 8: Đơn vị tính (ghi đè "Vỉ"->"Thùng" cho cột đầu; cột Bia hơi
    # không khai -> vẫn giữ đvt suy ra mặc định "Lít").
    ws.cell(row=7, column=6, value="CSPS330")
    ws.cell(row=7, column=7, value="CSPS330")
    ws.cell(row=8, column=6, value="Thùng")
    ws.cell(row=8, column=7, value="Lon")
    ws.cell(row=9, column=1, value="06631")
    ws.cell(row=9, column=2, value="Tr Hà")
    ws.cell(row=9, column=3, value="Nam Thu")
    ws.cell(row=9, column=6, value=208)
    ws.cell(row=9, column=7, value=16)
    ws.cell(row=9, column=8, value=5)
    ws.cell(row=10, column=1, value="TỔNG KEG")
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    result = _import(client, admin_h, content)
    assert len(result["HL"]) == 1
    slip = client.get(f"/api/wms/load-slips/{result['HL'][0]['load_slip_id']}", headers=admin_h).json()
    lines = {l["product_name"].strip(): l for l in slip["lines"]}

    assert lines["Vỉ SP Sleek"]["uom"] == "Thùng"  # ghi đè, không phải "Vỉ" suy ra từ tên cột
    assert lines["Vỉ SP Sleek"]["product_code"] == "CSPS330"  # dòng Mã sản phẩm vẫn hoạt động
    assert lines["LON SP Sleek KM"]["uom"] == "Lon"  # khai lại đúng giá trị mặc định vẫn nhận
    assert lines["Bia hơi 30L"]["uom"] == "Lít"  # cột không khai đvt -> vẫn suy ra như cũ


def test_uom_row_before_code_row_still_detected_correctly(client, admin_h):
    """Thứ tự 2 dòng khai báo có thể ngược lại: Đơn vị tính TRƯỚC, Mã sản phẩm SAU — vẫn phải
    phân loại đúng từng dòng theo nội dung, không phụ thuộc thứ tự cố định."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HL"
    ws["A1"] = "LỆNH ĐÓNG HÀNG HL"
    ws["A2"] = "Ca 3  -  Ngày 13 tháng 7 năm 2026"
    headers = ["SỐ XE", "TÊN LX", "NPP VÀ NVBH", "GHI CHÚ", "SỐ QĐ KM", "Vỉ SP Golden"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=6, column=c, value=h)
    ws.cell(row=7, column=6, value="Két")  # Đơn vị tính TRƯỚC
    ws.cell(row=8, column=6, value="CSGN330")  # Mã sản phẩm SAU
    ws.cell(row=9, column=1, value="06631")
    ws.cell(row=9, column=2, value="Tr Hà")
    ws.cell(row=9, column=3, value="Nam Thu")
    ws.cell(row=9, column=6, value=100)
    ws.cell(row=10, column=1, value="TỔNG KEG")
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    result = _import(client, admin_h, content)
    slip = client.get(f"/api/wms/load-slips/{result['HL'][0]['load_slip_id']}", headers=admin_h).json()
    line = slip["lines"][0]
    assert line["uom"] == "Két"
    assert line["product_code"] == "CSGN330"
    assert line["quantity"] == 100


def test_slip_code_format_and_year(client, admin_h):
    content = _build_workbook(date_text="Ngày 5 tháng 3 năm 2026")
    result = _import(client, admin_h, content)
    for v in result["HL"]:
        assert v["slip_code"].endswith("/2026/BBBG-BHL")


def test_list_update_header_and_delete(client, admin_h):
    content = _build_workbook()
    result = _import(client, admin_h, content)
    load_slip_id = result["HL"][0]["load_slip_id"]

    listed = client.get("/api/wms/load-slips?sheet_type=HL", headers=admin_h).json()
    assert any(s["load_slip_id"] == load_slip_id for s in listed)

    updated = client.put(f"/api/wms/load-slips/{load_slip_id}", headers=admin_h, json={
        "issuer_name": "Nguyễn Văn Tùng", "issuer_title": "Thủ kho",
        "recipient_title": "Lái xe",
    })
    assert updated.status_code == 200, updated.text
    assert updated.json()["issuer_name"] == "Nguyễn Văn Tùng"
    assert updated.json()["recipient_title"] == "Lái xe"
    assert updated.json()["recipient_name"]  # vẫn giữ giá trị mặc định (tên lái xe) đã có sẵn

    deleted = client.delete(f"/api/wms/load-slips/{load_slip_id}", headers=admin_h)
    assert deleted.status_code == 204, deleted.text
    missing = client.get(f"/api/wms/load-slips/{load_slip_id}", headers=admin_h)
    assert missing.status_code == 404


def test_import_requires_permission(client, admin_h):
    thukho_h = _login(client, "thukho", "123456")
    content = _build_workbook()
    r = client.post("/api/wms/load-slips/import", headers=thukho_h,
                    files={"file": ("x.xlsx", content,
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 201, r.text  # thukho có warehouse.issue


def _build_workbook_with_npp_code(plate, driver_text, npp_code, npp_text):
    """1 xe/1 dòng, có thêm cột 'MÃ NHÀ PHÂN PHỐI' — dùng để test ưu tiên tra Danh mục xe/NCC
    thay vì dùng thẳng chữ tự do ở cột TÊN LX/NPP VÀ NVBH."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HL"
    ws["A1"] = "LỆNH ĐÓNG HÀNG HL"
    ws["A2"] = "Ca 3  -  Ngày 20 tháng 8 năm 2026"
    headers = ["SỐ XE", "TÊN LX", "MÃ NHÀ PHÂN PHỐI", "NPP VÀ NVBH", "GHI CHÚ", "SỐ QĐ KM", "Vỉ Legend"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=6, column=c, value=h)
    ws.cell(row=7, column=1, value=plate)
    ws.cell(row=7, column=2, value=driver_text)
    ws.cell(row=7, column=3, value=npp_code)
    ws.cell(row=7, column=4, value=npp_text)
    ws.cell(row=7, column=7, value=20)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def test_npp_code_and_plate_resolve_from_catalog(client, admin_h):
    """Xe đã có trong Danh mục xe (Vehicle.plate) + mã khớp Danh mục Nhà cung cấp
    (Supplier.code) -> ưu tiên lấy TÊN LX/NPP từ 2 danh mục đó, KHÔNG dùng chữ tự do trong
    file (vốn cố tình ghi sai/khác để phân biệt rõ nguồn dữ liệu nào thắng)."""
    veh = client.post("/api/wms/vehicles", headers=admin_h,
                      json={"plate": "77777", "driver_name": "Tài xế Chuẩn (danh mục)"})
    assert veh.status_code == 201, veh.text
    sup = client.post("/api/suppliers", headers=admin_h,
                      json={"code": "NPP-CATALOG-01", "name": "Công ty TNHH Phân phối Chuẩn"})
    assert sup.status_code == 201, sup.text

    content = _build_workbook_with_npp_code("77777", "Tên sai trong file", "NPP-CATALOG-01",
                                            "Tên NPP sai trong file")
    result = _import(client, admin_h, content)
    slip_id = result["HL"][0]["load_slip_id"]
    slip = client.get(f"/api/wms/load-slips/{slip_id}", headers=admin_h).json()
    assert slip["driver_name"] == "Tài xế Chuẩn (danh mục)"
    assert slip["routes"] == "Công ty TNHH Phân phối Chuẩn"


def test_npp_code_and_plate_fallback_when_not_in_catalog(client, admin_h):
    """Xe/mã NPP KHÔNG có trong danh mục -> vẫn import được, dùng lại chữ tự do trong file
    (TÊN LX) làm phương án dự phòng; mã NPP không khớp thì hiện lại chính mã đó (không âm
    thầm bỏ qua, không tự ý ghép sang chữ tự do ở cột NPP VÀ NVBH để tránh nhầm lẫn)."""
    content = _build_workbook_with_npp_code("88888", "Tài xế Tự Do", "NPP-KHONG-TON-TAI",
                                            "Tên NPP tự do")
    result = _import(client, admin_h, content)
    slip_id = result["HL"][0]["load_slip_id"]
    slip = client.get(f"/api/wms/load-slips/{slip_id}", headers=admin_h).json()
    assert slip["driver_name"] == "Tài xế Tự Do"
    assert slip["routes"] == "NPP-KHONG-TON-TAI"
