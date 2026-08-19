"""Test "Lệnh đóng hàng" theo ngày (LoadOrder) — gộp các Biên bản bàn giao (LoadSlip) của
1 lần import Excel:
1) Import tạo đúng 1 LoadOrder/sheet có xe, gộp hết xe của sheet đó, mã số theo quy tắc
   "số lớn nhất đã dùng" (giống slip_code) với hậu tố riêng "/LDH-BHL".
2) GET chi tiết lệnh trả về đủ danh sách xe kèm lines (tái dùng đúng shape của get_load_slip).
3) "+ Thêm xe": chỉ liệt kê xe CHƯA thuộc lệnh nào, cùng sheet_type — xe đã gán vào lệnh (dù
   lệnh nào) biến mất khỏi danh sách khả dụng.
4) "Bỏ" xe khỏi lệnh trả xe đó về trạng thái chưa gán — xuất hiện lại trong danh sách khả dụng.
5) Không cho thêm xe khác sheet_type, không cho thêm xe đã thuộc lệnh khác."""

import io
import os
import tempfile

_TMP = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
os.environ["MES_DATABASE_URL"] = f"sqlite:///{_TMP.name}"
os.environ["MES_DEV_HEADER_AUTH"] = "0"
os.environ["MES_RL_ENABLED"] = "0"
os.environ["MES_ADMIN_PASSWORD"] = "AdminTest123"

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.main import app
from app import seed as seed_mod


@pytest.fixture(scope="module", autouse=True)
def _seeded():
    seed_mod.seed()
    yield


@pytest.fixture(scope="module")
def client():
    return TestClient(app)


def _login(client, u, p):
    r = client.post("/api/auth/login", json={"username": u, "password": p})
    assert r.status_code == 200, r.text
    return {"Authorization": "Bearer " + r.json()["token"]}


@pytest.fixture(scope="module")
def admin_h(client):
    return _login(client, "admin", "AdminTest123")


def _build_workbook(sheet_name, vehicles, date_text="Ngày 15 tháng 8 năm 2026"):
    """vehicles: list of (plate, driver, npp, qty_bia_hoi)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws["A1"] = f"LỆNH ĐÓNG HÀNG {sheet_name}"
    ws["A2"] = f"Ca 3  -  {date_text}"
    headers = ["SỐ XE", "TÊN LX", "NPP VÀ NVBH", "GHI CHÚ", "SỐ QĐ KM", "Bia hơi 30L"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=6, column=c, value=h)
    r = 7
    for plate, driver, npp, qty in vehicles:
        ws.cell(row=r, column=1, value=plate)
        ws.cell(row=r, column=2, value=driver)
        ws.cell(row=r, column=3, value=npp)
        ws.cell(row=r, column=6, value=qty)
        r += 1
    ws.cell(row=r, column=1, value="TỔNG KEG")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _import(client, admin_h, content, filename="lenh-dong-hang.xlsx"):
    r = client.post("/api/wms/load-slips/import", headers=admin_h,
                    files={"file": (filename, content,
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 201, r.text
    return r.json()


def test_import_creates_one_load_order_grouping_all_vehicles(client, admin_h):
    content = _build_workbook("HL", [
        ("11111", "Tài xế A", "NPP1", 10),
        ("22222", "Tài xế B", "NPP2", 5),
    ])
    result = _import(client, admin_h, content)
    assert len(result["HL"]) == 2
    assert "HL" in result["load_orders"]
    load_order_id = result["load_orders"]["HL"]["load_order_id"]
    order_code = result["load_orders"]["HL"]["order_code"]
    assert order_code.endswith("/2026/LDH-BHL")

    listed = client.get("/api/wms/load-orders?sheet_type=HL", headers=admin_h).json()
    assert any(o["load_order_id"] == load_order_id for o in listed)
    row = next(o for o in listed if o["load_order_id"] == load_order_id)
    assert row["vehicle_count"] == 2

    detail = client.get(f"/api/wms/load-orders/{load_order_id}", headers=admin_h).json()
    assert detail["order_code"] == order_code
    plates = {v["vehicle_plate"] for v in detail["vehicles"]}
    assert plates == {"11111", "22222"}
    v1 = next(v for v in detail["vehicles"] if v["vehicle_plate"] == "11111")
    assert v1["lines"][0]["product_name"].strip() == "Bia hơi 30L"
    assert v1["lines"][0]["quantity"] == 10


def test_two_separate_imports_get_sequential_order_codes(client, admin_h):
    c1 = _build_workbook("HL", [("33333", "X", "NPP", 1)])
    c2 = _build_workbook("HL", [("44444", "Y", "NPP", 1)])
    r1 = _import(client, admin_h, c1)
    r2 = _import(client, admin_h, c2)
    code1 = r1["load_orders"]["HL"]["order_code"]
    code2 = r2["load_orders"]["HL"]["order_code"]
    seq1 = int(code1.split("/")[0])
    seq2 = int(code2.split("/")[0])
    assert seq2 == seq1 + 1


def test_available_vehicles_excludes_already_assigned_and_add_remove_flow(client, admin_h):
    content = _build_workbook("ĐM", [
        ("55555", "Tài xế C", "NPP3", 7),
    ])
    result = _import(client, admin_h, content)
    load_order_id = result["load_orders"]["ĐM"]["load_order_id"]
    load_slip_id = result["ĐM"][0]["load_slip_id"]

    # Xe vừa import đã tự động nằm trong lệnh -> không còn xuất hiện ở danh sách khả dụng.
    avail = client.get(f"/api/wms/load-orders/{load_order_id}/available-vehicles", headers=admin_h).json()
    assert not any(v["load_slip_id"] == load_slip_id for v in avail)

    # Bỏ xe khỏi lệnh -> xuất hiện lại trong danh sách khả dụng.
    removed = client.delete(f"/api/wms/load-orders/{load_order_id}/vehicles/{load_slip_id}", headers=admin_h)
    assert removed.status_code == 200, removed.text
    assert removed.json()["vehicles"] == []
    avail2 = client.get(f"/api/wms/load-orders/{load_order_id}/available-vehicles", headers=admin_h).json()
    assert any(v["load_slip_id"] == load_slip_id for v in avail2)

    # Thêm lại vào đúng lệnh đó -> thành công, không còn ở danh sách khả dụng nữa.
    added = client.post(f"/api/wms/load-orders/{load_order_id}/vehicles", headers=admin_h,
                        json={"load_slip_id": load_slip_id})
    assert added.status_code == 201, added.text
    assert len(added.json()["vehicles"]) == 1
    avail3 = client.get(f"/api/wms/load-orders/{load_order_id}/available-vehicles", headers=admin_h).json()
    assert not any(v["load_slip_id"] == load_slip_id for v in avail3)

    # Thêm lần 2 (đã thuộc lệnh) -> lỗi rõ ràng, không cho trùng.
    dup = client.post(f"/api/wms/load-orders/{load_order_id}/vehicles", headers=admin_h,
                      json={"load_slip_id": load_slip_id})
    assert dup.status_code == 409, dup.text


def test_cannot_add_vehicle_from_different_sheet_type(client, admin_h):
    hl = _import(client, admin_h, _build_workbook("HL", [("66666", "X", "NPP", 1)]))
    dm = _import(client, admin_h, _build_workbook("ĐM", [("77777", "Y", "NPP", 1)]))
    hl_order_id = hl["load_orders"]["HL"]["load_order_id"]
    dm_slip_id = dm["ĐM"][0]["load_slip_id"]

    r = client.post(f"/api/wms/load-orders/{hl_order_id}/vehicles", headers=admin_h,
                    json={"load_slip_id": dm_slip_id})
    assert r.status_code == 409, r.text
