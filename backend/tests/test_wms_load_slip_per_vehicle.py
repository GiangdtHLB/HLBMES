"""Test luồng Xuất kho chọn theo TỪNG XE trong Lệnh đóng hàng (export_lines_for_slip) và khoá/mở
sửa dòng hàng theo trạng thái xuất (LoadSlip.shipment_id):
1) export_lines_for_slip trả đúng dữ liệu CHỈ của 1 xe (khác export_lines_for_order gộp cả đơn).
2) create_shipment(..., load_slip_id=...) gán đúng LoadSlip.shipment_id; gọi lại lần 2 với cùng
   load_slip_id -> lỗi rõ ràng (chặn double-submit).
3) update_load_slip_lines thành công khi shipment_id is None; lỗi khi đã có shipment_id.
4) undo_shipment xoá LoadSlip.shipment_id về None -> update_load_slip_lines lại thành công."""

import io
import os
import tempfile

_TMP = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
os.environ["MES_DATABASE_URL"] = f"sqlite:///{_TMP.name}"
os.environ["MES_DEV_HEADER_AUTH"] = "0"
os.environ["MES_RL_ENABLED"] = "0"
os.environ["MES_ADMIN_PASSWORD"] = "AdminTest123"

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.main import app
from app import seed as seed_mod


@pytest.fixture(scope="module", autouse=True)
def _seeded():
    seed_mod.seed()
    yield


@pytest.fixture(scope="module")
def client():
    return TestClient(app)


def _login(client, u, p):
    r = client.post("/api/auth/login", json={"username": u, "password": p})
    assert r.status_code == 200, r.text
    return {"Authorization": "Bearer " + r.json()["token"]}


@pytest.fixture(scope="module")
def admin_h(client):
    return _login(client, "admin", "AdminTest123")


def _build_workbook_2_vehicles():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HL"
    ws["A1"] = "LỆNH ĐÓNG HÀNG HL"
    ws["A2"] = "Ca 3  -  Ngày 19 tháng 8 năm 2026"
    headers = ["SỐ XE", "TÊN LX", "NPP VÀ NVBH", "GHI CHÚ", "SỐ QĐ KM", "Vỉ Legend"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=6, column=c, value=h)
    ws.cell(row=7, column=6, value="CLGN330PV")
    ws.cell(row=8, column=1, value="XE-A")
    ws.cell(row=8, column=2, value="Tài xế A")
    ws.cell(row=8, column=5, value="NPP-A")
    ws.cell(row=8, column=6, value=20)
    ws.cell(row=9, column=1, value="XE-B")
    ws.cell(row=9, column=2, value="Tài xế B")
    ws.cell(row=9, column=5, value="NPP-B")
    ws.cell(row=9, column=6, value=15)
    ws.cell(row=10, column=1, value="TỔNG KEG")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@pytest.fixture(scope="module")
def setup(client, admin_h):
    fp = client.post("/api/finished-products", headers=admin_h, json={
        "code": "CLGN330PV", "name": "Bia lon Legend 330ml (per-vehicle test)",
        "unit_type": "vi", "pack_size": 24,
    })
    assert fp.status_code == 201, fp.text
    fp_id = fp.json()["finished_product_id"]

    wh = client.post("/api/wms/warehouses", headers=admin_h,
                     json={"code": "WH-PV-HL", "name": "Kho Test PV Hạ Long",
                           "load_order_sheet_type": "HL"})
    assert wh.status_code == 201, wh.text
    warehouse_id = wh.json()["warehouse_id"]

    loc = client.post("/api/wms/locations", headers=admin_h,
                      json={"code": "LOC-PV-A1", "name": "Khu PV A1", "capacity": 1000,
                            "warehouse_id": warehouse_id})
    assert loc.status_code == 201, loc.text
    loc_id = loc.json()["loc_id"]

    ob = client.post("/api/wms/units", headers=admin_h,
                     json={"finished_product_id": fp_id, "product_name": "CLGN330PV",
                           "lot_code": "OB-PV-01", "total": 50 * 24, "pack_size": 24,
                           "unit_type": "vi", "is_opening_balance": True, "loc_id": loc_id})
    assert ob.status_code == 201, ob.text

    ship_to = client.post("/api/suppliers", headers=admin_h,
                          json={"code": "PV-SHIPTO", "name": "Test ship-to PV"})
    assert ship_to.status_code == 201, ship_to.text
    ship_to_id = ship_to.json()["supplier_id"]

    content = _build_workbook_2_vehicles()
    imp = client.post("/api/wms/load-slips/import", headers=admin_h,
                      files={"file": ("lenh-dong-hang-pv.xlsx", content,
                                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert imp.status_code == 201, imp.text
    load_order_id = imp.json()["load_orders"]["HL"]["load_order_id"]

    order = client.get(f"/api/wms/load-orders/{load_order_id}", headers=admin_h).json()
    vehicles = {v["vehicle_plate"]: v["load_slip_id"] for v in order["vehicles"]}

    return {"warehouse_id": warehouse_id, "ship_to_id": ship_to_id,
            "load_order_id": load_order_id, "slip_a": vehicles["XE-A"], "slip_b": vehicles["XE-B"]}


def test_list_load_slips_includes_shipment_status(client, admin_h, setup):
    # GET /wms/load-slips (bảng Sheet HL/ĐM) phải trả kèm shipment_id/shipment_code cho mọi xe,
    # kể cả khi shipment_ids rỗng (không có xe nào đã xuất) — regression cho lỗi thiếu .all()
    # trên Result khi dict(db.execute(...)) (phát hiện qua kiểm thử trình duyệt thủ công).
    r = client.get("/api/wms/load-slips", headers=admin_h)
    assert r.status_code == 200, r.text
    rows = r.json()
    assert any(row["load_slip_id"] == setup["slip_a"] for row in rows)


def test_export_lines_for_slip_scoped_to_one_vehicle(client, admin_h, setup):
    r = client.get(f"/api/wms/load-slips/{setup['slip_a']}/export-suggestion", headers=admin_h)
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["vehicle_plate"] == "XE-A"
    assert len(data["lines"]) == 1
    assert data["lines"][0]["quantity"] == 20
    assert data["warehouse_id"] == setup["warehouse_id"]

    r_b = client.get(f"/api/wms/load-slips/{setup['slip_b']}/export-suggestion", headers=admin_h)
    assert r_b.json()["lines"][0]["quantity"] == 15

    # Order-level vẫn gộp cả 2 xe (35), khác slip-level.
    r_order = client.get(f"/api/wms/load-orders/{setup['load_order_id']}/export-suggestion", headers=admin_h)
    assert r_order.json()["lines"][0]["quantity"] == 35


def test_shipment_locks_slip_and_blocks_reselect(client, admin_h, setup):
    slip_before = client.get(f"/api/wms/load-slips/{setup['slip_a']}", headers=admin_h).json()
    assert slip_before["shipment_id"] is None

    created = client.post("/api/wms/shipments", headers=admin_h, json={
        "ship_to_id": setup["ship_to_id"], "warehouse_id": setup["warehouse_id"],
        "load_slip_id": setup["slip_a"],
        "lines": [{"product_name": "CLGN330PV", "unit_type": "vi", "quantity": 20}],
    })
    assert created.status_code == 201, created.text
    shipment_id = created.json()["shipment_id"]

    slip_after = client.get(f"/api/wms/load-slips/{setup['slip_a']}", headers=admin_h).json()
    assert slip_after["shipment_id"] == shipment_id
    assert slip_after["shipment_code"] == created.json()["shipment_code"]

    # Xe B chưa xuất -> vẫn None, không bị ảnh hưởng.
    slip_b = client.get(f"/api/wms/load-slips/{setup['slip_b']}", headers=admin_h).json()
    assert slip_b["shipment_id"] is None

    # Chọn lại xe A lần 2 -> chặn (409 DomainError).
    dup = client.post("/api/wms/shipments", headers=admin_h, json={
        "ship_to_id": setup["ship_to_id"], "warehouse_id": setup["warehouse_id"],
        "load_slip_id": setup["slip_a"],
        "lines": [{"product_name": "CLGN330PV", "unit_type": "vi", "quantity": 5}],
    })
    assert dup.status_code == 409, dup.text

    setup["shipment_id"] = shipment_id


def test_update_lines_blocked_while_shipped_then_unblocked_after_undo(client, admin_h, setup):
    blocked = client.put(f"/api/wms/load-slips/{setup['slip_a']}/lines", headers=admin_h,
                         json={"lines": [{"product_name": "CLGN330PV", "uom": "Vỉ", "quantity": 25}]})
    assert blocked.status_code == 409, blocked.text

    undone = client.post(f"/api/wms/shipments/{setup['shipment_id']}/undo", headers=admin_h)
    assert undone.status_code == 200, undone.text

    slip_after_undo = client.get(f"/api/wms/load-slips/{setup['slip_a']}", headers=admin_h).json()
    assert slip_after_undo["shipment_id"] is None

    saved = client.put(f"/api/wms/load-slips/{setup['slip_a']}/lines", headers=admin_h,
                       json={"lines": [{"product_name": "CLGN330PV", "uom": "Vỉ", "quantity": 25,
                                        "product_code": "CLGN330PV"}]})
    assert saved.status_code == 200, saved.text
    assert len(saved.json()["lines"]) == 1
    assert saved.json()["lines"][0]["quantity"] == 25
